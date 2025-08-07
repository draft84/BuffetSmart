from django.db.models import Q, F, Avg, ExpressionWrapper, FloatField, Count, Value, Prefetch
from django.conf import settings
from django.contrib.auth.models import User as DjangoUser
from django.core.cache import cache
from django.core.paginator import Paginator
from .models import UserApp, Organization, Dishes, DishesLang, Language, Recipe, Allergens, AllergensLang, Assignment, MenuManagement, Time, Week, Exception, DayliMenu, Logs, Role
from labels.models import Label, Template, Restaurant
import requests
import json
import pytz
import logging
import time
import aiohttp
import base64
from django.db import connection
from openpyxl import Workbook
from collections import OrderedDict
from googletrans import Translator
from django.utils.timezone import make_aware
from django.utils import timezone
# from dishes.logger_config import setup_logger
from logging.handlers import TimedRotatingFileHandler
from datetime import date, datetime, timedelta
import tzlocal
from django.http import JsonResponse
from dishes.schema import set_schema
from contextlib import contextmanager
from django_tenants.utils import schema_context
from django.db import connection
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import deepl
from django.contrib.auth.models import User
from django.db import transaction
import pandas as pd
from django.http import HttpResponse
import io
from django.db.models import Case, When, Max, Value, BooleanField, CharField
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from .models import Assignment
from rest_framework import status
from .models import Assignment
from rest_framework import status

@contextmanager
def schema_context(user):
    schema_name = set_schema(user)  # Configura el esquema
    try:
        yield schema_name  # Continúa la operación
    finally:
        # Restaura al esquema por defecto
        connection.cursor().execute(f'SET search_path TO public;')

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Configura el logger para crear un log de acciones realizadas en la aplicacion
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def setup_logger():
    fecha_actual = date.today().strftime('%Y-%m-%d')
    
    nombre_archivo = f"sincroni_log_{fecha_actual}.log"

    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    manejador_archivo = TimedRotatingFileHandler(
        nombre_archivo, when='D', interval=1, backupCount=30)
    manejador_archivo.setFormatter(logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s'))

    logger.addHandler(manejador_archivo)
    return logger


# Ejemplo de uso
logger = setup_logger()

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Extraemos la compañia y el codigo del store
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def store_code(user):

    set_schema(user)

    global company
    global code

    # Obtén el primer (y único) registro de la tabla Organization
    store = Organization.objects.first()

    if store:
        company = store.company
        code = store.store_code
    else:
        print("Las variables company estan vacias")

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Generamos el token para acceder al api de solum
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def solum_token_generate(user):
    try:
        set_schema(user)
        store_code(user)
        print(f"El usuario para solicitar el token es {user}")

        # Verifica que 'company' esté definido
        if not company:
            logger.error("El valor de 'company' no está definido.")
            return None, None

        token_url = f"https://eu.common.solumesl.com/common/api/v2/token?company={company}"
        token_payload = json.dumps({
            "username": "artemi@sincroni.io",
            "password": "@urora%esl2024"
        })
        token_headers = {
            'Content-Type': 'application/json'
        }

        logger.info(f"Enviando solicitud a {token_url}")
        token_response = requests.post(
            token_url, headers=token_headers, data=token_payload)
        token_response.raise_for_status()

        data = token_response.json()
        logger.info("Token generado exitosamente")
        return company, data
    except requests.exceptions.RequestException as e:
        logger.error(f"Error en la solicitud HTTP: {e}")
        return None, None
    except json.JSONDecodeError as e:
        logger.error(f"Error al decodificar la respuesta JSON: {e}")
        return None, None
    except Exception as e:
        logger.error(f"Error inesperado: {e}")
        return None, None

def refresh_solum_token(user, refresh_token, company):
    try:
        # Supongamos que la API externa dispone de un endpoint para refrescar el token
        refresh_url = f"https://eu.common.solumesl.com/common/api/v2/refresh?company={company}"
        payload = json.dumps({
            "refresh_token": refresh_token
        })
        headers = {'Content-Type': 'application/json'}
        logger.info(f"Enviando solicitud de refresh a {refresh_url}")
        response = requests.post(refresh_url, headers=headers, data=payload)
        response.raise_for_status()
        data = response.json()
        logger.info("Token refrescado exitosamente")
        return data
    except requests.exceptions.RequestException as e:
        logger.error(f"Error en la solicitud de refresh HTTP: {e}")
        return None
    except json.JSONDecodeError as e:
        logger.error(
            f"Error al decodificar la respuesta JSON del refresh: {e}")
        return None
    except Exception as e:
        logger.error(f"Error inesperado en el refresh: {e}")
        return None

def get_valid_solum_token(user):
    """
    Obtiene un token válido: si existe en caché y no ha expirado, se usa; de lo
    contrario, se intenta refrescarlo o se genera uno nuevo.
    """
    token_cache_key = f"solum_token_{user.id}"
    token_data = cache.get(token_cache_key)

    if token_data:
        # token_data debe contener: access_token, refresh_token, expires_at, company
        if timezone.now() < token_data['expires_at']:
            # El access token aún es válido
            return token_data
        else:
            # Token expirado: se intenta refrescar
            new_data = refresh_solum_token(
                user, token_data['refresh_token'], token_data['company'])
            if new_data and 'responseMessage' in new_data and 'access_token' in new_data['responseMessage']:
                response = new_data['responseMessage']
                access_token = response['access_token']
                refresh_token = response['refresh_token']
                # Valor por defecto en segundos
                expires_in = response.get('expires_in', 86400)
                expires_at = timezone.now() + datetime.timedelta(seconds=expires_in)
                new_token_data = {
                    'access_token': access_token,
                    'refresh_token': refresh_token,
                    'expires_at': expires_at,
                    'company': token_data['company']
                }
                cache.set(token_cache_key, new_token_data, expires_in)
                return new_token_data
            else:
                # Si falla el refresco, se solicita un token nuevo
                company, data = solum_token_generate(user)
                if data and 'responseMessage' in data and 'access_token' in data['responseMessage']:
                    response = data['responseMessage']
                    access_token = response['access_token']
                    refresh_token = response['refresh_token']
                    expires_in = response.get('expires_in', 86400)
                    expires_at = timezone.now() + datetime.timedelta(seconds=expires_in)
                    token_data = {
                        'access_token': access_token,
                        'refresh_token': refresh_token,
                        'expires_at': expires_at,
                        'company': company
                    }
                    cache.set(token_cache_key, token_data, expires_in)
                    return token_data
    else:
        # No hay token en caché; se solicita uno nuevo
        company, data = solum_token_generate(user)
        if data and 'responseMessage' in data and 'access_token' in data['responseMessage']:
            response = data['responseMessage']
            access_token = response['access_token']
            refresh_token = response['refresh_token']
            expires_in = response.get('expires_in', 86400)
            expires_at = timezone.now() + timedelta(seconds=expires_in)
            token_data = {
                'access_token': access_token,
                'refresh_token': refresh_token,
                'expires_at': expires_at,
                'company': company
            }
            cache.set(token_cache_key, token_data, expires_in)
            return token_data

    return None
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Crea la lista de articulos y asignaciones en solum con el total de labels que tiene su nomenclatura ----------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def install_product_labels(user):

    store_code(user)

    # Nombre del esquema
    schema_name, schema_id, schema_role = set_schema(user)

    assigmentList = []
    messages = []
    products = []

    description = None

    template = None

    # Se obtiene un token válido (se refresca o genera uno nuevo si fuera necesario)
    token_data = get_valid_solum_token(user)
    if not token_data:
        messages.append("No se pudo obtener un token válido")
        return messages

    # Extrae el access_token y company desde token_data
    access_token = token_data['access_token']

    labels = Label.objects.filter(~Q(label=F('mac')))

    # Obtenemos el numero total de productos y labels a cagar
    total_install = labels.count()

    for a in labels:

        # Crea la lista fija de productos
        # -----------------------------------------------
        # -----------------------------------------------
        products.append({
            "stationCode": f"{code}",
            "id": f"{a.mac}",
            "name": f"{a.label}",
            "nfc": f"https://www.solumesl.com/p/{a.label}",
            "eans": [],
            "data": {
                "STORE_CODE": f"{code}",
                "ITEM_ID": f"{a.mac}",
                "ITEM_NAME": f"{a.label}",
                "IS_EMPTY": "YES",
                "HAS_EXCEPTION": "NO",
            }
        })

        assigmentList.append({"articleIdList": [f"{a.mac}"], "labelCode": f"{a.mac}", "templateName": a.template.name})

    createProductsAndLabels = create_products_and_lables(company, products, access_token, user)

    if createProductsAndLabels == 200 and assigmentList != []:

        createAssignments, response_message = create_assignments(company, code, assigmentList, access_token, user)

        if createAssignments == 200:

            print(createAssignments)

            messages.append(f"Productos y labels creados: {total_install}")
            messages.append(f"Asignaciones creadas: {total_install}")
            messages.append(f"Instalacion creada en el esquema: {schema_name}")
            messages.append(f"Template seleccionado para el esquema Template: {schema_name} es {template}")

            logging.info(f"-------------------------------------------------------------------")
            logging.info(f"-------------------------------------------------------------------")
            logging.info(f"Se ha ejecutado la instalacion de productos y asignacion de labesl")
            logging.info(f"Productos: {total_install}")
            logging.info(f"Asignaciones creadas: {total_install}")
            logging.info(f"-------------------------------------------------------------------")
            logging.info(f"-------------------------------------------------------------------")

            description = f"Productos y asignaciones creadas: {total_install}"

            # Guardamos en la tabla de logs
            # -----------------------------------------------
            # -----------------------------------------------
            Logs.objects.create(function="create_assignments", description=description, status="Success")

        if createAssignments == 403 or createAssignments == 405:

            Logs.objects.create(function="create_assignments", description=f"Error al crear asignacion de productos revisar datos a cargar o labels ({response_message})", status="Error")
            messages.append(f"{response_message}")
    else:
        messages.append("Fallo la instalacion de productos en Solum no cargaron los productos")
        logging.info("Fallo la instalacion de productos en Solum no cargaron los productos")
        Logs.objects.create(function="create_products_and_lables", description="Fallo la instalacion de productos en Solum no cargaron los productos", status="Error")
        
    print(messages)

    return messages

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Listar template ----------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def load_labels_template(user):
    # Obtener el nombre del esquema
    schema_name, schema_id, schema_role = set_schema(user)

    templates = None
    templateList = []

    # bluesea Comprobar si el esquema contiene 'bs_'
    if schema_name.startswith('bs_'):
        templates = Template.objects.filter(name__icontains='BLUESEA')
        # template = '42_BLUESEA_V2_4_LANG'

    # barcelo Comprobar si el esquema contiene 'bf_'
    if schema_name.startswith('bf_'):
        templates = Template.objects.filter(name__icontains='BARCELO')
        # template = '42_Barcelo_V5'

    # kimpton Comprobar si el esquema contiene 'kp_'
    if schema_name.startswith('kp_'):
        templates = Template.objects.filter(name__icontains='KIMPTON')
        # template = '42_IHG_V1_3_LANG'

    # lopesan Comprobar si el esquema contiene 'lp_'
    if schema_name.startswith('lp_'):
        templates = Template.objects.filter(name__icontains='LOPESAN')
        # template = '42_LOPESAN_V1_3_LANG'
    
    # monument Comprobar si el esquema contiene 'monu'
    if schema_name.startswith('monu'):
        templates = Template.objects.filter(name__icontains='35_BARCELO_V2')
        # template = '42_LOPESAN_V1_3_LANG'

    for t in templates:
        templateList.append({"id": t.id, "templateName": t.name})

    return json.dumps(templateList)

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Listar template Crea la lista de articulos y asignaciones en solum con el total de labels que tiene su nomenclatura ------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def assignment_template(user, template, mac=None):

    logging.info(f"Template asignado: {template}")

    store_code(user)
    
    # Nombre del esquema
    schema_name, schema_id, schema_role = set_schema(user)

    # Se obtiene un token válido (se refresca o genera uno nuevo si fuera necesario)
    token_data = get_valid_solum_token(user)
    if not token_data:
        messages.append("No se pudo obtener un token válido")
        return messages

    # Extrae el access_token y company desde token_data
    access_token = token_data['access_token']

    assigmentList = []
    messages = []
    products = []

    description = None

    if mac:
        
        labels = Label.objects.filter(mac=mac).update(template_id=template)
        
        t = Template.objects.filter(id=template).first()
        
        logging.info(f"El template seleccionado es: {t.name}")
        
        l = Label.objects.filter(mac=mac).first()
        
        assigmentList.append({"articleIdList": [f"{l.mac}"], "labelCode": f"{l.mac}", "templateName": t.name})
        
    else:
        t = Template.objects.filter(id=template).first()

        logging.info(f"El template seleccionado es: {t.name}")

        logging.info(f"{access_token}")

        labels = Label.objects.filter(~Q(label=F('mac')))

        for a in labels:

            assigmentList.append({"articleIdList": [f"{a.mac}"], "labelCode": f"{a.mac}", "templateName": t.name})

    logging.info(f"{assigmentList}")

    createAssignments = create_assignments(company, code, assigmentList, access_token, user)
    
    print(createAssignments)

    # [0] para tomar el primer valor de la tupla que devuelve 200, SUCCESS
    if createAssignments[0] == 200:

        l = Label.objects.all().update(template_id=template)

        messages.append(f"Se ha cargado con exito el template {template}")

        description = f"Se ha cargado con exito el template {template}"

        # Guardamos en la tabla de logs
        # -----------------------------------------------
        # -----------------------------------------------
        Logs.objects.create(function="assignment_template", description=description, status="Success")

    if createAssignments == 405:
        messages.append(f"Error al asignar el template {template}")
        Logs.objects.create(function="assignment_template", description='Error al asignar el template verifique', status="Error")

    return messages

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Cargar articulos en solum ------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def create_products_and_lables(company, products, access_token, user):

    set_schema(user)

    messages = []

    # URL para registrar el artículo
    article_url = f"https://eu.common.solumesl.com/common/api/v1/articles?company={company}"

    # Datos para la solicitud de registro de artículo
    article_payload = json.dumps({
        "dataList": products
    }
    )

    article_headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {access_token}'
    }

    # Solicitud para registrar el artículo
    article_response = requests.post(
        article_url, headers=article_headers, data=article_payload)

    return article_response.status_code

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Cargar asignaciones en solum ---------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def create_assignments(company, code, assigmentList, access_token, user):
    
    set_schema(user)

    # Construir la URL y el payload para la solicitud
    labels_url = f"https://eu.common.solumesl.com/common/api/v2/common/labels/link?company={company}&store={code}"
    labels_payload = json.dumps({"assignList": assigmentList})

    labels_headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {access_token}'
    }

    # Realizar la solicitud POST a la API
    labels_response = requests.request("POST", labels_url, headers=labels_headers, data=labels_payload)

    # Inicializar response_message en None
    response_message = None

    # Intentar parsear la respuesta a JSON para extraer el responseMessage
    try:
        response_data = labels_response.json()
        response_message = response_data.get("responseMessage")
    except Exception as e:
        print("Error al procesar la respuesta JSON:", e)
        
    print(labels_response.status_code)

    # Retornar el status code y el responseMessage
    return labels_response.status_code, response_message

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Cargamos la lista de gateway disponibles en un store
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def mac_adderss_list(user):
    # Configuramos el esquema y obtenemos los valores de store
    set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']
    company = token_data['company']

    # Se construye la URL usando el company y el código de store
    mac_address_url = f"https://eu.common.solumesl.com/common/api/v2/common/gateway?company={company}&store={code}"

    mac_address_headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {access_token}'
    }

    try:
        mac_address_response = requests.get(
            mac_address_url, headers=mac_address_headers)
        # Lanza una excepción si el estado no es 200
        mac_address_response.raise_for_status()

        mac_data = mac_address_response.json()
        message = []

        if 'gatewayList' in mac_data:
            for gateway in mac_data['gatewayList']:
                mac = gateway.get('macAddress')
                if mac:
                    # Imprime (o registra) la dirección MAC obtenida y la agrega al mensaje
                    print(mac)
                    message.append(mac)
        return message
    except requests.exceptions.RequestException as e:
        logger.error(f"Error en la llamada a mac-address-list: {e}")
        return []

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Sincronizamos la lista de labes del gateway seleccionado
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def sync_labels_of_solum(mac_address, user):

    # Configuramos el esquema y obtenemos los valores de store
    set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']
    company = token_data['company']

    solum_url = f"https://eu.common.solumesl.com/common/api/v2/common/labels/gateway?company={company}&store={code}&gateway={mac_address}&network=true&sort=labelCode%2Casc&&size=500&page=0"

    solum_headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {access_token}'
    }

    solum_response = requests.get(solum_url, headers=solum_headers)

    if solum_response.status_code == 200:

        logging.info(
            f"sync_labels_of_solum => {solum_response.status_code}")

        solum_data = solum_response.json()
        if 'labelList' in solum_data:
            label_list = solum_data['labelList']
            for label in label_list:
                labelCode = label.get('labelCode')
                labelModel = label.get('labelModel')

                if not Label.objects.filter(mac=labelCode).exists():
                    # Crear o actualizar el Label solo si no existe una plantilla con el mismo mac
                    Label.objects.create(
                        mac=labelCode,
                        label=labelCode,
                        model=labelModel,
                        enabled='False'
                    )

                    description = f"Sincronizados lables del gateway {mac_address}"

                    # Guardamos en la tabla de logs
                    # -----------------------------------------------
                    # -----------------------------------------------
                    Logs.objects.create(
                        function="sync_labels_of_solum", description=description, status="Success")
                    messages = description
                else:
                    logging.info(
                        f"sync_labels_of_solum => El Label {labelCode} ya existe y no se guarda nuevamente.")
                    messages = "Los labels ya existen y no se volveran a cargar"
        else:
            description = f"No se encontraron lables del gateway {mac_address}"
            Logs.objects.create(
                function="sync_labels_of_solum", description=description, status="Error")
            messages = description
    else:
        description = "Error de conexion con SOLUM"
        Logs.objects.create(function="sync_labels_of_solum", description = description, status = "Error")
        messages = description

    return messages

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Cargamos la actualizacion de templates del api solum
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def sync_template_of_solum(user):

    # Configuramos el esquema y obtenemos los valores de store
    set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']
    
    messages=[]
    
    company = token_data['company']
    solum_url = f"https://eu.common.solumesl.com/common/api/v2/common/templates?company={company}"

    solum_payload = {}

    solum_headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {access_token}'
    }

    solum_response = requests.get(solum_url, headers=solum_headers, data=solum_payload)

    if solum_response.status_code == 200:

        logging.info(f"sync_template_of_solum => Respuesta del api code: {solum_response.status_code}")

        templates = solum_response.json().get('templateList', [])

        for template in templates:
            template_name = template.get('templateName')
            width = template.get('width')
            height = template.get('height')
            label_type = f"{width} x {height}"

            if template_name:
                print(f"Procesando plantilla: {template_name}")

                logging.info(f"Procesando plantilla: {template_name}")

                preview_response = execute_template_preview(template_name, access_token, user)
                if 'previewResponse' in preview_response:
                    preview_items = preview_response['previewResponse']
                    for item in preview_items:
                        if 'content' in item:
                            base_64_image = item['content']
                            save_template_name = template_name.replace('_COMPANY_SIN.xsl', '')
                            save_template_to_db(save_template_name, label_type, base_64_image, user)
                            messages.append(f"Plantilla guardada: {save_template_name}, label: {label_type}, imagen base 64: {base_64_image}")
                            # logging.info(f"sync_template_of_solum => Plantilla guardada: {save_template_name}, label: {label_type}")
                        else:
                            messages.append(f"Error al obtener 'content' para la plantilla: {template_name}")
                else:
                    print(f"Error al obtener la previsualización para la plantilla: {template_name}")

        messages.append("Sincronización exitosa.")
        f"Plantilla guardada: {save_template_name}, label: {label_type}, imagen base 64: {base_64_image}"
        logging.info(f"sync_template_of_solum => Plantilla guardada: {save_template_name}, label: {label_type}")
    else:
        messages.append(f"Fallo en obtener las plantillas: {solum_response.status_code} - {solum_response.text}")
        logging.error(f"sync_template_of_solum => Fallo en obtener las plantillas: {solum_response.status_code} - {solum_response.text}")
    

    return messages

def execute_template_preview(template_name, access_token, user):

    set_schema(user)

    url = f"https://eu.common.solumesl.com/common/api/v2/common/preview/template?company=SIN"

    payload = json.dumps({
        "template": template_name,
        "previewArticles": [
            {
                "index": 1,
                "articleId": "100001",
                "articleName": "APPLE1S",
                "data": {
                    "SALE_PRICE": "100",
                    "DISCOUNT_PRICE": "10"
                }
            }
        ]
    })
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {access_token}'
    }

    response = requests.post(url, headers=headers, data=payload)

    if response.status_code == 200:
        print('template', template_name)
        return response.json()


    else:

        print(f"Error al ejecutar la previsualización: {response.status_code} - {response.text}")

        return {"error": "Failed to execute preview", "status_code": response.status_code}

def save_template_to_db(save_template_name, label_type, base_64_image, user):

    try:
        set_schema(user)

        if not Template.objects.filter(name=save_template_name, label_type=label_type).exists():

            Template.objects.create(name=save_template_name, label_type=label_type, base_64_image=base_64_image)
            print(f"Plantilla guardada: {save_template_name}")
            logging.info(f"sync_template_of_solum => Plantilla guardada: {save_template_name}")

        else:
            # Si ya existe, no hacer nada
            print(f"La plantilla {save_template_name} con tipo {label_type} ya existe. No se guarda nuevamente.")
            logging.info(f"sync_template_of_solum => La plantilla {save_template_name} con tipo {label_type} ya existe. No se guarda nuevamente.")

    except:

        print(f"Error al guardar la plantilla {save_template_name}")

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Correr la rueda con las asignaciones del dia -----------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# def labels_template(user, restaurant=None):
#     """
#     Carga las etiquetas (labels) para el usuario y un restaurante opcional.
#     Retorna una lista de mensajes de diagnóstico.
#     """
#     messages = []
#     try:
#         # -------------------------------
#         # 1. CONFIGURACIÓN INICIAL
#         # -------------------------------
#         schema_name, schema_id, schema_role = set_schema(user)
#         store_code(user)  # Asigna variables globales 'company' y 'code'

#         token_data = get_valid_solum_token(user)
#         if not token_data:
#             logging.error("No se pudo obtener un token válido")
#             messages.append("Error: no se pudo obtener token de Solum.")
#             return messages

#         access_token = token_data.get('access_token')

#         # -------------------------------
#         # 2. ZONA HORARIA Y FECHA/HORA
#         # -------------------------------
#         tz, fecha_actual, dia_semana, hora_actual_dt = hour_zone(user)
#         hora_actual_dt = _normalize_to_time(hora_actual_dt)
#         time.sleep(1)
#         Logs.objects.create(
#             function="hour_zone",
#             description="Obtener la zona horaria asignada en SOLUM",
#             status="Success"
#         )

#         # -------------------------------
#         # 3. OBTENER SEMANAS CORTA Y LARGA
#         # -------------------------------
#         week_short_qs = Week.objects.filter(
#             start_date__lte=fecha_actual,
#             end_date__gte=fecha_actual,
#             active=True
#         ).annotate(
#             rango_dias=F('end_date') - F('start_date')
#         ).filter(rango_dias__lt=datetime.timedelta(days=365))

#         week_long_qs = Week.objects.filter(
#             start_date__lte=fecha_actual,
#             end_date__gte=fecha_actual,
#             active=True
#         ).annotate(
#             rango_dias=F('end_date') - F('start_date')
#         ).filter(rango_dias__gte=datetime.timedelta(days=365))

#         week_short_ids = list(week_short_qs.values_list('id', flat=True))
#         week_long_ids = list(week_long_qs.values_list('id', flat=True))

#         assignments = None
#         used_week_ids = []

#         # ----------------------------------------------------------
#         # 4. FIJAR QUERIES DE ASIGNACIONES SEGÚN CORTA/LARGA Y RESTAURANTE
#         # ----------------------------------------------------------

#         # ---- RAMA A: existe semana corta Y hay RESTAURANTE ----
#         if week_short_ids and restaurant is not None:
#             messages.append(f"[CASO A] Existe semana corta & restaurant={restaurant} → probando con IDs: {week_short_ids}")
#             qs_short = Assignment.objects.filter(
#                 day_of_week=dia_semana,
#                 time__start__lte=hora_actual_dt,
#                 time__end__gte=hora_actual_dt,
#                 label__restaurant_id=restaurant,
#                 week_id__in=week_short_ids
#             ).select_related('time', 'label').prefetch_related(
#                 Prefetch('exception_set', queryset=Exception.objects.all(), to_attr='exceptions')
#             )
#             if qs_short.exists():
#                 assignments = qs_short
#                 used_week_ids = week_short_ids
#             else:
#                 messages.append(f"[CASO A1] No encontró en semana corta con restaurant. Ahora probando semana larga IDs: {week_long_ids}")
#                 qs_long = Assignment.objects.filter(
#                     day_of_week=dia_semana,
#                     time__start__lte=hora_actual_dt,
#                     time__end__gte=hora_actual_dt,
#                     label__restaurant_id=restaurant,
#                     week_id__in=week_long_ids
#                 ).select_related('time', 'label').prefetch_related(
#                     Prefetch('exception_set', queryset=Exception.objects.all(), to_attr='exceptions')
#                 )
#                 assignments = qs_long
#                 used_week_ids = week_long_ids

#         # ---- RAMA B: existe semana corta Y NO hay RESTAURANTE ----
#         elif week_short_ids and restaurant is None:
#             messages.append(f"[CASO B] Existe semana corta & restaurant=None → probando con IDs: {week_short_ids}")
#             qs_short = Assignment.objects.filter(
#                 day_of_week=dia_semana,
#                 time__start__lte=hora_actual_dt,
#                 time__end__gte=hora_actual_dt,
#                 week_id__in=week_short_ids
#             ).select_related('time').prefetch_related(
#                 Prefetch('exception_set', queryset=Exception.objects.all(), to_attr='exceptions')
#             )
#             if qs_short.exists():
#                 assignments = qs_short
#                 used_week_ids = week_short_ids
#             else:
#                 messages.append(f"[CASO B1] No encontró en semana corta sin restaurant. Ahora probando semana larga IDs: {week_long_ids}")
#                 qs_long = Assignment.objects.filter(
#                     day_of_week=dia_semana,
#                     time__start__lte=hora_actual_dt,
#                     time__end__gte=hora_actual_dt,
#                     week_id__in=week_long_ids
#                 ).select_related('time').prefetch_related(
#                     Prefetch('exception_set', queryset=Exception.objects.all(), to_attr='exceptions')
#                 )
#                 assignments = qs_long
#                 used_week_ids = week_long_ids

#         # ---- RAMA C: NO existe semana corta Y hay RESTAURANTE ----
#         elif not week_short_ids and restaurant is not None:
#             messages.append(f"[CASO C] No existe semana corta & restaurant={restaurant} → usando semana larga IDs: {week_long_ids}")
#             qs_long = Assignment.objects.filter(
#                 day_of_week=dia_semana,
#                 time__start__lte=hora_actual_dt,
#                 time__end__gte=hora_actual_dt,
#                 label__restaurant_id=restaurant,
#                 week_id__in=week_long_ids
#             ).select_related('time', 'label').prefetch_related(
#                 Prefetch('exception_set', queryset=Exception.objects.all(), to_attr='exceptions')
#             )
#             assignments = qs_long
#             used_week_ids = week_long_ids

#         # ---- RAMA D: NO existe semana corta Y NO hay RESTAURANTE ----
#         else:  # equivale a “not week_short_ids and restaurant is None”
#             messages.append(f"[CASO D] No existe semana corta & restaurant=None → usando semana larga IDs: {week_long_ids}")
#             qs_long = Assignment.objects.filter(
#                 day_of_week=dia_semana,
#                 time__start__lte=hora_actual_dt,
#                 time__end__gte=hora_actual_dt,
#                 week_id__in=week_long_ids
#             ).select_related('time').prefetch_related(
#                 Prefetch('exception_set', queryset=Exception.objects.all(), to_attr='exceptions')
#             )
#             assignments = qs_long
#             used_week_ids = week_long_ids

#         total_assignments = assignments.count()
#         messages.append(f"Total asignaciones encontradas: {total_assignments} (usando week_ids = {used_week_ids})")

#         if total_assignments == 0:
#             # No hay asignaciones ni en la semana corta ni en la larga
#             messages.extend([
#                 "No hay asignaciones válidas en la franja horaria programada.",
#                 f"Fecha: {fecha_actual}, Día de la semana: {dia_semana}, Hora: {hora_actual_dt}"
#             ])
#             Logs.objects.create(
#                 function="labels_template",
#                 description="Franja horaria sin asignaciones válidas (ni en semana corta ni en larga).",
#                 status="Success"
#             )
#             return messages

#         # -------------------------------
#         # 5. PROCESAR ASIGNACIONES Y EXCEPCIONES
#         # -------------------------------
#         assignment_list_payload = []
#         for asign in assignments:
#             # Procesar la asignación principal (sin excepción)
#             _process_assignment_or_exception(
#                 assignment=asign,
#                 is_exception=False,
#                 exception_obj=None,
#                 payload_list=assignment_list_payload,
#                 base_item={
#                     "articleId": asign.label.mac,
#                     "articleName": asign.label.label
#                 },
#                 schema_name=schema_name
#             )

#             # Si hay excepciones, procesar cada una
#             for exc in asign.exceptions:
#                 _process_assignment_or_exception(
#                     assignment=asign,
#                     is_exception=True,
#                     exception_obj=exc,
#                     payload_list=assignment_list_payload,
#                     base_item={
#                         "articleId": Label.objects.get(id=exc.label_id).mac,
#                         "articleName": Label.objects.get(id=exc.label_id).label
#                     },
#                     schema_name=schema_name
#                 )

#         # -------------------------------
#         # 6. ENVIAR DATOS A SOLUM
#         # -------------------------------
#         company = globals().get('company')
#         code = globals().get('code')
#         eupdate_url = (
#             f"https://eu.common.solumesl.com/common/api/v2/common/articles?"
#             f"company={company}&store={code}"
#         )
#         eupdate_headers = {
#             'Content-Type': 'application/json',
#             'Authorization': f'Bearer {access_token}'
#         }
#         eupdate_payload = json.dumps(assignment_list_payload)

#         try:
#             eupdate_resp = requests.put(
#                 eupdate_url,
#                 headers=eupdate_headers,
#                 data=eupdate_payload,
#                 timeout=10
#             )
#             eupdate_resp.raise_for_status()
#         except Exception as e:
#             logging.error(f"Error al actualizar artículos en Solum: {e}")
#             messages.append("Error: no se pudo actualizar los artículos en Solum.")
#             return messages

#         time.sleep(1)

#         # -------------------------------
#         # 7. REGISTRAR EN DAYLI MENU Y LOGS
#         # -------------------------------
#         asign_first = assignments.first()
#         exception_qs = Exception.objects.filter(
#             date=fecha_actual,
#             time__start__lte=hora_actual_dt,
#             time__end__gte=hora_actual_dt,
#             week_id__in=used_week_ids
#         )
#         total_exceptions = exception_qs.count()

#         current_date = datetime.date.today()
#         turn_name = asign_first.time.name
#         if not DayliMenu.objects.filter(date=current_date, turn=turn_name).exists():
#             DayliMenu.objects.create(
#                 assignments=total_assignments,
#                 turn=turn_name,
#                 hour_start=asign_first.time.start,
#                 hour_end=asign_first.time.end,
#                 exception=total_exceptions
#             )
#         else:
#             DayliMenu.objects.filter(date=current_date, turn=turn_name).update(
#                 exception=total_exceptions
#             )

#         description_log = (
#             f"Se cargaron {total_assignments} asignaciones en turno '{turn_name}', "
#             f"Company: {company}, Code: {code}, Fecha: {fecha_actual} {hora_actual_dt}, "
#             f"Excepciones: {total_exceptions}"
#         )
#         Logs.objects.create(
#             function="labels_template",
#             description=description_log,
#             status="Success"
#         )
#         messages.append(description_log)

#         # -------------------------------
#         # 8. ACTUALIZAR ARTICLES
#         # -------------------------------
#         update_articles(company, code, assignment_list_payload, access_token)

#         return messages

#     except Exception as err:
#         logging.exception(f"Error inesperado en labels_template: {err}")
#         messages.append(f"Error inesperado: {err}")
#         return messages


# # ----------------------
# # FUNCIONES AUXILIARES
# # ----------------------

# def _normalize_to_time(hora_actual_dt):
#     """
#     Asegura que 'hora_actual_dt' sea siempre un datetime.time.
#     Si es string, intenta formatos ISO o 'HH:MM:SS'.
#     """
#     if isinstance(hora_actual_dt, datetime.datetime):
#         return hora_actual_dt.time()
#     if isinstance(hora_actual_dt, str):
#         try:
#             return datetime.datetime.fromisoformat(hora_actual_dt).time()
#         except ValueError:
#             return datetime.datetime.strptime(hora_actual_dt, "%H:%M:%S").time()
#     if isinstance(hora_actual_dt, datetime.time):
#         return hora_actual_dt
#     raise ValueError(f"Tipo de dato no válido para hora: {type(hora_actual_dt)}")


# def _calculate_allergens_and_vegetarian(dish_obj):
#     """
#     Dado un objeto Dish, recorre sus Recipe y .allergens para:
#       - determinar si es vegetariano (si existe un allergen.id == 0)
#       - construir un dict con los contadores de alérgenos distintos
#     Retorna: (vegetarian_flag, dict_alergenos)
#     """
#     alergenos_all = {
#         "ALTRAMUCES": "", "APIO": "", "CACAHUETES": "", "CRUSTACEO": "",
#         "FRUTOS_SECOS": "", "GLUTEN": "", "HUEVOS": "", "LACTEOS": "",
#         "MOLUSCO": "", "MOSTAZA": "", "PESCADO": "", "SESAMO": "",
#         "SOJA": "", "SULFITO": ""
#     }
#     vegetarian = "NO"
#     contador = 0

#     recipes = Recipe.objects.filter(dish=dish_obj).prefetch_related('allergens')
#     for receta in recipes:
#         for allergen in receta.allergens.all():
#             if allergen.id == 0:
#                 vegetarian = "YES"
#             else:
#                 contador += 1
#                 alergenos_all[allergen.allergen.upper()] = str(contador)
#                 vegetarian = "NO"

#     return vegetarian, alergenos_all


# def _get_translations_dict(dish_id, schema_name):
#     """
#     Construye un dict de traducciones para el dish_id en todos los idiomas activos,
#     siguiendo regla especial si schema_name comienza con 'kp_' (conversion a minus+capitalizado sólo en FR).
#     Retorna un diccionario con claves "TRANSLATION_1", "TRANSLATION_2", ...
#     """
#     translation_base = {
#         "TRANSLATION_1": "", "TRANSLATION_2": "",
#         "TRANSLATION_3": "", "TRANSLATION_4": ""
#     }

#     languages = Language.objects.filter(status=True).order_by('position')
#     translations_dict = translation_base.copy()

#     for idx, lang in enumerate(languages, start=1):
#         entry = DishesLang.objects.filter(dish_id=dish_id, language=lang.code).first()
#         if not entry or not entry.translation:
#             continue

#         text = entry.translation
#         if schema_name.startswith('kp_') and lang.code == "FR":
#             text = text.lower().capitalize()
#         translations_dict[f"TRANSLATION_{idx}"] = text

#     return translations_dict


# def _process_assignment_or_exception(assignment, is_exception, exception_obj, payload_list, base_item, schema_name):
#     """
#     Dado un objeto Assignment o Exception, construye el payload JSON que irá en payload_list.
#       - Si is_exception=False, toma datos de 'assignment' (assignment.dish + assignment.label).
#       - Si is_exception=True, toma datos de 'exception_obj' (exception_obj.dish_id + exception_obj.label_id).
#     Al final agrega un diccionario completo a 'payload_list'.
#     """
#     if is_exception:
#         dish_id = exception_obj.dish_id
#         label_obj = Label.objects.get(id=exception_obj.label_id)
#         dishlang_list = list(DishesLang.objects.filter(dish_id=dish_id).values_list('translation', flat=True))
#     else:
#         dish_id = assignment.dish.id
#         label_obj = assignment.label
#         dishlang_list = []

#     # Obtener objeto Dish (o None si dish_id == 0)
#     dish_obj = Dishes.objects.filter(id=dish_id).first() if dish_id != 0 else None

#     # Calcular alérgenos y vegetariano
#     if dish_obj:
#         vegetariano_flag, alergenos_dict = _calculate_allergens_and_vegetarian(dish_obj)
#     else:
#         # Si dish_id == 0, usamos valores por defecto
#         vegetariano_flag = "YES"
#         alergenos_dict = {k: "" for k in _calculate_allergens_and_vegetarian(dish_obj)[1].keys()}

#     # Construir traducciones (o valores por defecto si dish_id == 0)
#     if dish_obj:
#         translations_dict = _get_translations_dict(dish_id, schema_name)
#     else:
#         translations_dict = {
#             "TRANSLATION_1": "DISH",
#             "TRANSLATION_2": "GERICHT",
#             "TRANSLATION_3": "PLAT",
#             "TRANSLATION_4": ""
#         }

#     # Construir los campos comunes de 'data'
#     item_description = "PLATO" if dish_id == 0 else assignment.dish.dish.upper()
#     data_common = {
#         "ITEM_DESCRIPTION": item_description,
#         "IS_EMPTY": "YES" if dish_id == 0 else "NO",
#         "HAS_EXCEPTION": "NO" if not is_exception else "YES",
#         "IS_VEGETARIAN": vegetariano_flag
#     }

#     if is_exception:
#         data_common.update({
#             "EXCEPTION_NAME": dish_obj.dish if dish_obj else "",
#             "EXCEPTION_DESCRIPTION": ", ".join(dishlang_list),
#             "EXCEPTION_CATEGORY_03": ""
#         })
#     else:
#         data_common.update({
#             "EXCEPTION_NAME": "",
#             "EXCEPTION_DESCRIPTION": "",
#             "EXCEPTION_CATEGORY_03": ""
#         })

#     # Unir alérgenos y traducciones
#     data_common.update(alergenos_dict)
#     data_common.update(translations_dict)

#     # Construir el dict final para el artículo
#     article_payload = {
#         "articleId": base_item["articleId"],
#         "articleName": base_item["articleName"],
#         "nfcUrl": "",
#         "eans": [],
#         "data": data_common
#     }

#     payload_list.append(article_payload)



# def labels_template(user, restaurant=None):
    
#     # Configuramos el esquema y obtenemos los valores de store
#     schema_name, schema_id, schema_role = set_schema(user)
#     store_code(user)  # Esto define las variables globales "company" y "code"

#     # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
#     token_data = get_valid_solum_token(user)
#     if not token_data:
#         logger.error("No se pudo obtener un token válido")
#         return []

#     # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
#     access_token = token_data['access_token']

#     count = 0
#     allCount = 0
#     messages = []
#     assigmentList = []
#     description = None
#     assigmentList.clear()    
#     vegetarian = "NO"
    
#     alergenos_all = {"ALTRAMUCES": "", "APIO": "", "CACAHUETES": "", "CRUSTACEO": "", "FRUTOS_SECOS": "", "GLUTEN": "", "HUEVOS": "", "LACTEOS": "", "MOLUSCO": "", "MOSTAZA": "", "PESCADO": "", "SESAMO": "", "SOJA": "", "SULFITO": ""}
#     translation_all = {"TRANSLATION_1": "", "TRANSLATION_2": "", "TRANSLATION_3": "", "TRANSLATION_4": ""}
    
#     timezone, fecha_actual, dia_semana, hora_actual_dt = hour_zone(user)
    
#     # — Normalizar hora_actual_dt para que sea siempre datetime.time —
#     if isinstance(hora_actual_dt, str):
#         # intenta ISO datetime "YYYY-MM-DDTHH:MM:SS[.ffffff]" o "HH:MM:SS" 
#         try:
#             # para cadena completa con fecha y hora
#             hora_actual_dt = datetime.datetime.fromisoformat(hora_actual_dt).time()
#         except ValueError:
#             # para cadena solo hora
#             hora_actual_dt = datetime.datetime.strptime(hora_actual_dt, "%H:%M:%S").time()
    
#     elif isinstance(hora_actual_dt, datetime.datetime):
#         hora_actual_dt = hora_actual_dt.time()
#     # si ya es datetime.time, queda igual
#     time.sleep(1)

#     Logs.objects.create(function = "hour_zone", description = "Obtener la zona horaria asignada en SOLUM", status = "Success")

#     # ------------------------------------------------------------------------------------------
#     # ------------------------------------------------------------------------------------------
#     # Buscamos las asignaciones
#     # ------------------------------------------------------------------------------------------
#     # ------------------------------------------------------------------------------------------
        
#     # Verificar si existe una semana que contenga la fecha actual            
#     existe_fecha = Week.objects.filter(start_date__lte=fecha_actual, end_date__gte=fecha_actual, active=True).annotate(rango_dias=F('end_date') - F('start_date')).filter(rango_dias__lt=datetime.timedelta(days=365))
#     valid_weeks = Week.objects.filter(start_date__lte=fecha_actual, end_date__gte=fecha_actual, active=True).annotate(rango_dias=F('end_date') - F('start_date')).filter(rango_dias__gte=timedelta(days=365))


#     # Querys a realizar si la fecha o el restaurant existe o no
#     if existe_fecha and restaurant is not None:     
#         week_ids = existe_fecha.values_list('id', flat=True)
#         messages.append(f"El weekID es: {week_ids}")
        
#         # assignments = Assignment.objects.filter(day_of_week=dia_semana,label__restaurant_id=restaurant,time__start__lte=hora_actual_dt,time__end__gte=hora_actual_dt,week_id__in=week_ids)
#         # assignments_first = assignments.first()
        
#         qs_exceptions = Exception.objects.all()
#         assignments = (
#             Assignment.objects
#               # filtros sobre Assignment
#               .filter(
#                   day_of_week=dia_semana,
#                   time__start__lte=hora_actual_dt,
#                   time__end__gte=hora_actual_dt,
#                   label__restaurant_id=restaurant,
#                   week_id__in=week_ids
#               )
#               .select_related('time', 'label')
#               .prefetch_related(
#                   Prefetch(
#                       'exception_set',
#                       queryset=qs_exceptions,
#                       to_attr='exceptions'
#                   )
#               )
#         )
#         assignments_first = assignments.first()
        
        
#         # Si no hay asignaciones
#         if not assignments.exists():
#             week_ids = valid_weeks.values_list('id', flat=True)
#             messages.append(f"El weekID es: {week_ids}")
            
#             # assignments = Assignment.objects.filter(day_of_week=dia_semana, label__restaurant_id=restaurant, time__start__lte=hora_actual_dt, time__end__gte=hora_actual_dt,week_id__in=week_ids)
#             # assignments_first = assignments.first()
            
#             qs_exceptions = Exception.objects.all()
#             assignments = (
#                 Assignment.objects
#                 # filtros sobre Assignment
#                 .filter(
#                     day_of_week=dia_semana,
#                     time__start__lte=hora_actual_dt,
#                     time__end__gte=hora_actual_dt,
#                     label__restaurant_id=restaurant,
#                     week_id__in=week_ids
#                 )
#                 .select_related('time', 'label')
#                 .prefetch_related(
#                     Prefetch(
#                         'exception_set',
#                         queryset=qs_exceptions,
#                         to_attr='exceptions'
#                     )
#                 )
#             )
#             assignments_first = assignments.first()

#     if existe_fecha and restaurant is None:
#         week_ids = existe_fecha.values_list('id', flat=True)
#         messages.append(f"El weekID es: {week_ids}")
        
#         # assignments = Assignment.objects.filter(day_of_week=dia_semana,time__start__lte=hora_actual_dt,time__end__gte=hora_actual_dt,week_id__in=week_ids)
#         # assignments_first = assignments.first()
        
#         qs_exceptions = Exception.objects.all()
#         assignments = (Assignment.objects.filter(day_of_week=dia_semana,time__start__lte=hora_actual_dt,time__end__gte=hora_actual_dt,week_id__in=week_ids)
#             .select_related('time')
#             .prefetch_related(
#                 Prefetch(
#                     'exception_set',
#                     queryset=qs_exceptions,
#                     to_attr='exceptions'
#                 )
#             )
#         )
#         assignments_first = assignments.first()
        
#         # Si no hay asignaciones
#         if not assignments.exists():
#             week_ids = valid_weeks.values_list('id', flat=True)
#             messages.append(f"El weekID es: {week_ids}")
        
#             # assignments = Assignment.objects.filter(day_of_week=dia_semana, time__start__lte=hora_actual_dt, time__end__gte=hora_actual_dt,week_id__in=week_ids)
#             # assignments_first = assignments.first()
            
#             qs_exceptions = Exception.objects.all()
#             assignments = (
#                 Assignment.objects
#                 .filter(day_of_week=dia_semana, time__start__lte=hora_actual_dt, time__end__gte=hora_actual_dt, week_id__in=week_ids)
#                 .select_related('time')
#                 .prefetch_related(
#                     Prefetch(
#                         'exception_set',
#                         queryset=qs_exceptions,
#                         to_attr='exceptions'
#                     )
#                 )
#             )
#             assignments_first = assignments.first()

#     if not existe_fecha and restaurant is not None:
#         week_ids = valid_weeks.values_list('id', flat=True)
#         messages.append(f"El weekID es: {week_ids}")
        
#         # assignments = Assignment.objects.filter(day_of_week=dia_semana, label__restaurant_id=restaurant, time__start__lte=hora_actual_dt, time__end__gte=hora_actual_dt,week_id__in=week_ids)
#         # assignments_first = assignments.first()
#         qs_exceptions = Exception.objects.all()

#         assignments = (
#             Assignment.objects
#               # filtros sobre Assignment
#               .filter(
#                   day_of_week=dia_semana,
#                   time__start__lte=hora_actual_dt,
#                   time__end__gte=hora_actual_dt,
#                   label__restaurant_id=restaurant,    # filtro extra por restaurant_id
#                   week_id__in=week_ids
#               )
#               # traer el objeto Time y Label para evitar consultas adicionales
#               .select_related('time', 'label')
#               # precargar excepciones relacionadas
#               .prefetch_related(
#                   Prefetch(
#                       'exception_set',
#                       queryset=qs_exceptions,
#                       to_attr='exceptions'
#                   )
#               )
#         )
#         assignments_first = assignments.first()

#     if not existe_fecha and restaurant is None:
#         week_ids = valid_weeks.values_list('id', flat=True)
#         messages.append(f"El weekID es: {week_ids}")
        
#         # assignments = Assignment.objects.filter(day_of_week=dia_semana, time__start__lte=hora_actual_dt, time__end__gte=hora_actual_dt,week_id__in=week_ids)
#         # assignments_first = assignments.first()
        
#         # queryset de excepciones (sin filtrar, pero puedes aplicarle filtros adicionales si lo deseas)
#         print(f"dia_semana: {dia_semana}")
#         qs_exceptions = Exception.objects.all()
#         assignments = (
#             Assignment.objects
#             .filter(day_of_week=dia_semana, time__start__lte=hora_actual_dt, time__end__gte=hora_actual_dt, week_id__in=week_ids)
#             .select_related('time')
#             .prefetch_related(
#                 Prefetch(
#                     'exception_set',
#                     queryset=qs_exceptions,
#                     to_attr='exceptions'
#                 )
#             )
#         )
#         assignments_first = assignments.first()

#     # Obtenemos el numero total de registros a cagar
#     total_assignments = assignments.count()
    
#     # ------------------------------------------------------------------------------------------
#     # ------------------------------------------------------------------------------------------

#     if total_assignments!=0:
            
#         for a in assignments:
            
#             if a.time.start <= hora_actual_dt <= a.time.end:
                        
#                 count += 1
#                 dish_id=a.dish.id                                 
                
#                 translations = DishesLang.objects.filter(dish=a.id).values_list('translation', flat=True)
                    
#                 itemDescription = f"{a.dish.dish}"
                            
#                 dish = Dishes.objects.get(id=dish_id)
                    
#                 recipes = Recipe.objects.filter(dish=dish).prefetch_related('allergens')

#                 alergenos = alergenos_all.copy()
                
#                 for recipe in recipes:
#                     for allergen in recipe.allergens.all():
#                         if allergen.id == 0:
#                             vegetarian = "YES"
#                         else:
#                             allCount += 1
#                             alergenos[allergen.allergen.upper()] = str(allCount)
#                             vegetarian = "NO"          
                
#                 # --------------------------------------------------------------------------
#                 # Consulta para obtener idiomas con estado verdadero y posición ascendente
#                 # -------------------------------------------------------------------------- 
#                 translations_dict = translation_all.copy()
                    
#                 dish_id = a.dish_id

#                 # Paso 4: Obtener los idiomas activos ordenados por posición
#                 languages = Language.objects.filter(status=True).order_by('position')
                
#                 if schema_name.startswith('kp_'):

#                     # Paso 5: Para cada idioma, obtener la traducción correspondiente del plato
#                     for index, language in enumerate(languages, start=1):
#                         translation = DishesLang.objects.filter(dish_id=dish_id, language=language.code).first()
                        
#                         if translation and translation.translation:
                            
#                                 # Si el idioma es español ('es'), formateamos la traducción
#                             if language.code == "FR":
#                                 translation_text = translation.translation.lower().capitalize()
#                             else:
#                                 translation_text = translation.translation
                            
#                             # Almacenamos la traducción en el diccionario
#                             translations_dict[f"TRANSLATION_{index}"] = translation_text
#                 else:
                    
#                     # Paso 5: Para cada idioma, obtener la traducción correspondiente del plato
#                     for index, language in enumerate(languages, start=1):
#                         translation = DishesLang.objects.filter(dish_id=dish_id, language=language.code).first()
                        
#                         if translation:
#                             # Agregar la traducción al diccionario con el formato adecuado
#                             translations_dict[f"TRANSLATION_{index}"] = translation.translation

                            
#                 allCount=0
                
#                 logging.info(f"articleId: {a.label.mac}, articleName: {a.label.label}, VEGETARIANO : vegetarian, ITEM_ID: {a.label.mac},  ITEM_DESCRIPTION: {itemDescription.upper()}")
                
#                 if a.exceptions:
#                     for e in a.exceptions:
                        
#                         dish_id=e.dish_id
#                         assignment_id=a.id
#                         label_id=e.label_id

#                         # obtenemos el ID del plato
#                         dish=Dishes.objects.filter(id=dish_id).first()

#                         # Obtenemos el id del articulo a editar
#                         assignment=Assignment.objects.filter(id=assignment_id).first()

#                         # Obtenemos el plato original para agregar la excepcion
#                         original_dish=Dishes.objects.filter(id=a.dish_id).first()

#                         # Obtenemos las traducciones del plato
#                         dishlang=DishesLang.objects.filter(dish_id=dish_id).values_list('translation', flat=True)

#                         # Obtenemos el label del articulo a editar
#                         labels=Label.objects.filter(id=label_id).first()


#                         dish = Dishes.objects.get(id=dish_id)
#                         recipes = Recipe.objects.filter(dish=dish).prefetch_related('allergens')

#                         # Luego, modifica tu código actual así:
#                         alergenos = alergenos_all.copy()  # Crea una copia del diccionario base

#                         for recipe in recipes:
#                             for allergen in recipe.allergens.all():
#                                 if allergen.id == 0:
#                                     vegetarian = "YES"
#                                 else:
#                                     allCount += 1
#                                     alergenos[allergen.allergen.upper()] = str(allCount)
#                                     vegetarian = "NO"

#                         allCount=0

#                         # --------------------------------------------------------------------------
#                         # Consulta para obtener idiomas con estado verdadero y posición ascendente
#                         # --------------------------------------------------------------------------

#                         translations_dict = translation_all.copy()  # Crea una copia del diccionario base

#                         dish_id = e.dish_id

#                         # Paso 4: Obtener los idiomas activos ordenados por posición
#                         languages = Language.objects.filter(status=True).order_by('position')

#                         # Paso 5: Para cada idioma, obtener la traducción correspondiente del plato
#                         for index, language in enumerate(languages, start=1):
#                             translation = DishesLang.objects.filter(dish_id=dish_id, language=language.code).first()

#                             if translation:
#                                 # Agregar la traducción al diccionario con el formato adecuado
#                                 translations_dict[f"TRANSLATION_{index}"] = translation.translation
                        
#                         if dish_id == 0:
#                             assigmentList.append({
#                                 "articleId": f"{a.label.mac}",
#                                 "articleName": f"{a.label.label}",
#                                 "data": {
#                                 "ITEM_DESCRIPTION": "PLATO",
#                                 "IS_EMPTY": "YES",
#                                 "HAS_EXCEPTION": "NO",
#                                 "VEGETARIANO" : vegetarian,
#                                 "EXCEPTION_NAME":"",
#                                 "EXCEPTION_DESCRIPTION":"",
#                                 "EXCEPTION_CATEGORY_03":"",
#                                 "ALTRAMUCES": "",
#                                 "APIO": "",
#                                 "CACAHUETES": "",
#                                 "CRUSTACEO": "",
#                                 "FRUTOS_SECOS": "",
#                                 "GLUTEN": "",
#                                 "HUEVOS": "",
#                                 "LACTEOS": "",
#                                 "MOLUSCO": "",
#                                 "MOSTAZA": "",
#                                 "PESCADO": "",
#                                 "SESAMO": "",
#                                 "SOJA": "",
#                                 "SULFITO": "",
#                                 "TRANSLATION_1" : "DISH",
#                                 "TRANSLATION_2" : "GERICHT",
#                                 "TRANSLATION_3" : "PLAT",
#                                 "TRANSLATION_4" : ""
#                             }})
#                         else:
#                             assigmentList.append({
#                             "articleId": f"{labels.mac}",
#                             "articleName": f"{labels.label}",
#                             "nfcUrl": f"",
#                             "eans": [],
#                             "data": {
#                                 "ITEM_ID": f"{labels.mac}",
#                                 "ITEM_NAME": f"{labels.label}",
#                                 "ITEM_DESCRIPTION": f"{itemDescription.upper()}",
#                                 "IS_EMPTY": "NO",
#                                 "HAS_EXCEPTION": "YES",
#                                 "IS_VEGETARIAN": vegetarian,
#                                 "EXCEPTION_NAME":f"{dish.dish}",
#                                 "EXCEPTION_DESCRIPTION":", ".join(dishlang),
#                                 "EXCEPTION_CATEGORY_03": "",
#                                 **alergenos,
#                                 **translations_dict
#                                 }
#                             })
#                 else:        
#                     if dish_id == 0:
#                         assigmentList.append({
#                             "articleId": f"{a.label.mac}",
#                             "articleName": f"{a.label.label}",
#                             "data": {
#                             "ITEM_DESCRIPTION": "PLATO",
#                             "IS_EMPTY": "YES",
#                             "HAS_EXCEPTION": "NO",
#                             "VEGETARIANO" : vegetarian,
#                             "EXCEPTION_NAME":"",
#                             "EXCEPTION_DESCRIPTION":"",
#                             "EXCEPTION_CATEGORY_03":"",
#                             "ALTRAMUCES": "",
#                             "APIO": "",
#                             "CACAHUETES": "",
#                             "CRUSTACEO": "",
#                             "FRUTOS_SECOS": "",
#                             "GLUTEN": "",
#                             "HUEVOS": "",
#                             "LACTEOS": "",
#                             "MOLUSCO": "",
#                             "MOSTAZA": "",
#                             "PESCADO": "",
#                             "SESAMO": "",
#                             "SOJA": "",
#                             "SULFITO": "",
#                             "TRANSLATION_1" : "DISH",
#                             "TRANSLATION_2" : "GERICHT",
#                             "TRANSLATION_3" : "PLAT",
#                             "TRANSLATION_4" : ""
#                         }})
#                     else:
#                         assigmentList.append({
#                             "articleId": f"{a.label.mac}",
#                             "articleName": f"{a.label.label}",
#                             "nfcUrl": f"",
#                             "eans": [],
#                             "data": {
#                                 "ITEM_ID": f"{a.label.mac}",
#                                 "ITEM_NAME": f"{a.label.label}",
#                                 "ITEM_DESCRIPTION": f"{itemDescription.upper()}",
#                                 "IS_EMPTY": "NO",
#                                 "HAS_EXCEPTION": "NO",
#                                 "VEGETARIANO" : vegetarian,
#                                 "EXCEPTION_NAME": "",
#                                 "EXCEPTION_DESCRIPTION": "",
#                                 "EXCEPTION_CATEGORY_03": "",
#                                 **alergenos,
#                                 **translations_dict
#                             }
#                         })
        
#         eupdateUrl = f"https://eu.common.solumesl.com/common/api/v2/common/articles?company={company}&store={code}"

#         eupdatePayload = json.dumps(assigmentList)

#         eupdateHeaders = {
#             'Content-Type': 'application/json',
#             'Authorization': f'Bearer {access_token}'
#         }

#         eupdateResponse = requests.put(eupdateUrl, headers=eupdateHeaders, data=eupdatePayload)
#         time.sleep(1) 
                
#         if existe_fecha:  
#             assignments =assignments_first
#             exception = Exception.objects.filter(date=fecha_actual, time__start__lte=hora_actual_dt, time__end__gte=hora_actual_dt, week_id__in=week_ids)
#         else:
#             assignments = assignments_first
#             exception = Exception.objects.filter(date=fecha_actual, time__start__lte=hora_actual_dt, time__end__gte=hora_actual_dt,week_id__in=week_ids)
                
#         # Obtenemos el numero total de registros a cagar
#         total_exception = exception.count()
            
#         # ------------------------------------------------------------------------------------------
#         # ------------------------------------------------------------------------------------------
#         # Salvar log en tabla
#         # ------------------------------------------------------------------------------------------
#         # ------------------------------------------------------------------------------------------
            
#         current_date = datetime.date.today()
#         turn_name = assignments.time.name

#         if not DayliMenu.objects.filter(date=current_date, turn=turn_name).exists():
#             DayliMenu.objects.create(assignments=total_assignments, turn=assignments.time.name, hour_start=assignments.time.start, hour_end=assignments.time.end, exception=total_exception)
#         else:
#             DayliMenu.objects.filter(date=current_date, turn=turn_name).update(exception=total_exception)
            
#         description = f"Se cargaron {total_assignments}, en el turno {assignments.time.name}, Company: {company}, Code: {code}, Fecha: {fecha_actual} - {hora_actual_dt}, Excepciones para esta rueda: {total_exception}"
            
#         messages.append(f"Company: {company}, Code: {code}, Fecha: {fecha_actual} - {hora_actual_dt}, Total de asignaciones: {total_assignments}, Excepciones para esta rueda: {total_exception}")
        
#         Logs.objects.create(function = "labels_template", description = description, status = "Success")
                
#         # Validamos si el contador es mayor a cero y si lo es que coincida con el total de registros encontrados
#         if count > 0 and count == total_assignments and assigmentList !=[]:
#             time.sleep(1)           
#             updateArticles=update_articles(company, code, assigmentList, access_token)

#     else:
#         messages.append(f"Debes estar en una franja horaria programada para cargar asignaciones o tener asignaciones validas")
#         messages.append(f"Fecha: {fecha_actual}, Dia de la semana: {dia_semana}, Hora: {hora_actual_dt}")
            
#         description = "Debes estar en una franja horaria programada para cargar asignaciones o tener asignaciones validas"
#         Logs.objects.create(function = "labels_template", description = description, status = "Success")

#     return messages

def labels_template(user, restaurant=None):
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # Correr la rueda con las asignaciones del dia ------------------------------------------------------------------------------------------------------------------
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    
    # Configuramos el esquema y obtenemos los valores de store
    schema_name, schema_id, schema_role = set_schema(user)
    
    store_code(user)

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']

    count = 0
    allCount = 0
    messages = []
    bulk_values = []
    assigmentList = []
    description = None
    assigmentList.clear()    
    vegetarian = "NO"
    
    alergenos_all = {"ALTRAMUCES": "", "APIO": "", "CACAHUETES": "", "CRUSTACEO": "", "FRUTOS_SECOS": "", "GLUTEN": "", "HUEVOS": "", "LACTEOS": "", "MOLUSCO": "", "MOSTAZA": "", "PESCADO": "", "SESAMO": "", "SOJA": "", "SULFITO": ""}
    
    translation_all = {"TRANSLATION_1": "", "TRANSLATION_2": "", "TRANSLATION_3": "", "TRANSLATION_4": ""}
    
    timezone, fecha_actual, dia_semana, hora_actual_dt = hour_zone(user)
    
    # — Normalizar hora_actual_dt para que sea siempre datetime.time —
    if isinstance(hora_actual_dt, str):
        # intenta ISO datetime "YYYY-MM-DDTHH:MM:SS[.ffffff]" o "HH:MM:SS" 
        try:
            # para cadena completa con fecha y hora
            hora_actual_dt = datetime.fromisoformat(hora_actual_dt).time()
        except ValueError:
            # para cadena solo hora
            hora_actual_dt = datetime.strptime(hora_actual_dt, "%H:%M:%S").time()
    
    elif isinstance(hora_actual_dt, datetime):
        hora_actual_dt = hora_actual_dt.time()
    # si ya es datetime.time, queda igual
    time.sleep(1)
    
    EXISTS_DATE_BOOL = f"""SELECT EXISTS(SELECT 1 FROM {schema_name}.dishes_week WHERE start_date <= %s AND end_date   >= %s AND (end_date - start_date) < 365 AND active = TRUE) AS existe;""".format(schema_name)

    with connection.cursor() as cursor:
        cursor.execute(EXISTS_DATE_BOOL, [fecha_actual, fecha_actual])
        existe_fecha = cursor.fetchone()[0]

    if existe_fecha and restaurant is not None:
        sql = f"""
            SELECT
            a.*,
            t.start,
            t.end,
            l.mac,
            l.label,
            d.dish,
            l.restaurant_id,
            COUNT(e.id) OVER () AS excepciones,
            COALESCE(e.id::text, '')       AS exception_id,
            COALESCE(e.date::text, '')     AS exception_date,
            COALESCE(e.time_id::text, '')  AS exception_time_id,
            COALESCE(e.label_id::text, '') AS exception_label_id
            FROM {schema_name}.dishes_assignment AS a
            JOIN {schema_name}.dishes_time   AS t ON a.time_id  = t.id
            JOIN {schema_name}.dishes_week   AS w ON a.week_id  = w.id
            JOIN {schema_name}.labels_label  AS l ON a.label_id = l.id
            JOIN {schema_name}.dishes_dishes AS d ON a.dish_id  = d.id
            LEFT JOIN {schema_name}.dishes_exception AS e 
                ON e.assignment_id = a.id
            WHERE
            a.day_of_week                   = %s
            AND t.start       <= %s
            AND t.end         >= %s
            AND (w.start_date + t.start)    <= %s
            AND (w.end_date   + t.end)      >= %s
            AND w.restaurant_id = %s
            AND (w.end_date - w.start_date) < 365;
        """
        
        params = [
            dia_semana,
            hora_actual_dt,
            hora_actual_dt,
            fecha_actual,
            fecha_actual,
            restaurant,
        ]
    
    if existe_fecha and restaurant is None:
        
        sql = f"""
            SELECT
            a.*,
            t.start,
            t.end,
            l.mac,
            l.label,
            d.dish,
            COUNT(e.id) OVER () AS excepciones,
            COALESCE(e.id::text, '')       AS exception_id,
            COALESCE(e.date::text, '')     AS exception_date,
            COALESCE(e.time_id::text, '')  AS exception_time_id,
            COALESCE(e.label_id::text, '') AS exception_label_id
            FROM {schema_name}.dishes_assignment AS a
            JOIN {schema_name}.dishes_time   AS t ON a.time_id  = t.id
            JOIN {schema_name}.dishes_week   AS w ON a.week_id  = w.id
            JOIN {schema_name}.labels_label  AS l ON a.label_id = l.id
            JOIN {schema_name}.dishes_dishes AS d ON a.dish_id  = d.id
            LEFT JOIN {schema_name}.dishes_exception AS e 
                ON e.assignment_id = a.id
            WHERE
            a.day_of_week                   = %s
            AND t.start       <= %s
            AND t.end         >= %s
            AND (w.start_date + t.start)    <= %s
            AND (w.end_date   + t.end)      >= %s
            AND (w.end_date - w.start_date) < 365;
        """
        params = [
            dia_semana,
            hora_actual_dt,
            hora_actual_dt,
            fecha_actual,
            fecha_actual,
        ]
        
    if not existe_fecha and restaurant is not None:
        sql = f"""
            SELECT
                a.*,
                t.start,
                t.end,
                l.mac,
                l.label,
                d.dish,
                l.restaurant_id        AS label_restaurant_id,
                w.restaurant_id        AS week_restaurant_id,
                COUNT(e.id) OVER () AS all_exceptions,
                COALESCE(e.id::text, '')   AS exception_id,
                COALESCE(e.date::text, '') AS exception_date,
                COALESCE(e.time_id::text, '')   AS exception_time_id,
                COALESCE(e.label_id::text, '')  AS exception_label_id
            FROM {schema_name}.dishes_assignment    AS a
            JOIN {schema_name}.dishes_time          AS t ON a.time_id = t.id
            JOIN {schema_name}.dishes_week          AS w ON a.week_id = w.id
            JOIN {schema_name}.labels_label         AS l ON a.label_id = l.id
            JOIN {schema_name}.dishes_dishes        AS d ON a.dish_id = d.id
            LEFT JOIN {schema_name}.dishes_exception AS e ON e.assignment_id = a.id
            WHERE
                a.day_of_week = %s
                AND t.start       <= %s::time
                AND t.end         >= %s::time
                AND (w.start_date + t.start)    <= %s::date
                AND (w.end_date   + t.end)      >= %s::date
                AND w.restaurant_id = %s
                AND (w.end_date - w.start_date) >= 365
        """
    
        hora_actual_str = hora_actual_dt.strftime("%H:%M:%S")
        params = [
            dia_semana,
            hora_actual_str,
            hora_actual_str,
            fecha_actual,
            fecha_actual,
            restaurant, 
        ]
        
    if not existe_fecha and restaurant is None:    
        sql = f"""
            SELECT 
            a.*,  
            t.start,  
            t.end,  
            l.mac,  
            l.label,  
            d.dish,
            l.restaurant_id        AS label_restaurant_id,
            w.restaurant_id        AS week_restaurant_id,
            COUNT(e.id) OVER () AS all_exceptions,  
            COALESCE(e.id::text, '')   AS exception_id,  
            COALESCE(e.date::text, '') AS exception_date,  
            COALESCE(e.time_id::text, '')   AS exception_time_id,  
            COALESCE(e.label_id::text, '')  AS exception_label_id  
            FROM {schema_name}.dishes_assignment    AS a  
            JOIN {schema_name}.dishes_time          AS t ON a.time_id = t.id  
            JOIN {schema_name}.dishes_week          AS w ON a.week_id = w.id  
            JOIN {schema_name}.labels_label         AS l ON a.label_id = l.id  
            JOIN {schema_name}.dishes_dishes        AS d ON a.dish_id = d.id  
            LEFT JOIN {schema_name}.dishes_exception AS e ON e.assignment_id = a.id  
            WHERE  
            a.day_of_week = %s  
            AND t.start <= %s::time  
            AND t.end   >= %s::time  
            AND (w.start_date + t.start) <= %s::date  
            AND (w.end_date   + t.end)   >= %s::date  
            AND (w.end_date - w.start_date) >= 365
        """
    
        hora_actual_str = hora_actual_dt.strftime("%H:%M:%S")
        params = [
            dia_semana,
            hora_actual_str,
            hora_actual_str,
            fecha_actual,
            fecha_actual,
        ]
        
        print(cursor.mogrify(sql, params).decode('utf-8'))
    
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()

    # 4) Procesa resultados sin que explote si no hay filas
    assignments = [dict(zip(columns, row)) for row in rows]
    total_assignments = len(assignments)
    total_exceptions = rows[0][columns.index("all_exceptions")]

    if total_assignments!=0:
        
        for a in assignments:
            id              = a['id']
            dish_id         = a['dish_id']
            dish            = a['dish']
            time_id         = a['time_id']
            start           = a['start']
            end             = a['end']
            label           = a['label']
            label_id        = a['label_id']
            mac             = a['mac']
            exception_id    = a['exception_id']
            week_rest_id   = a['week_restaurant_id']
            exception_date  = a['exception_date']
            
            bulk_values.append((label, week_rest_id))
              
            if start <= hora_actual_dt <= end:
                        
                count += 1                               
                
                translations = DishesLang.objects.filter(dish=id).values_list('translation', flat=True)
                    
                itemDescription = f"{dish}"
                            
                dish = Dishes.objects.get(id=dish_id)
                    
                recipes = Recipe.objects.filter(dish=dish).prefetch_related('allergens')

                alergenos = alergenos_all.copy()
                
                for recipe in recipes:
                    for allergen in recipe.allergens.all():
                        if allergen.id == 0:
                            vegetarian = "YES"
                        else:
                            allCount += 1
                            alergenos[allergen.allergen.upper()] = str(allCount)
                            vegetarian = "NO"          
                
                # --------------------------------------------------------------------------
                # Consulta para obtener idiomas con estado verdadero y posición ascendente
                # -------------------------------------------------------------------------- 
                translations_dict = translation_all.copy()

                # Paso 4: Obtener los idiomas activos ordenados por posición
                languages = Language.objects.filter(status=True).order_by('position')
                
                if schema_name.startswith('kp_'):

                    # Paso 5: Para cada idioma, obtener la traducción correspondiente del plato
                    for index, language in enumerate(languages, start=1):
                        translation = DishesLang.objects.filter(dish_id=dish_id, language=language.code).first()
                        
                        if translation and translation.translation:
                            
                                # Si el idioma es español ('es'), formateamos la traducción
                            if language.code == "FR":
                                translation_text = translation.translation.lower().capitalize()
                            else:
                                translation_text = translation.translation
                            
                            # Almacenamos la traducción en el diccionario
                            translations_dict[f"TRANSLATION_{index}"] = translation_text
                else:
                    
                    # Paso 5: Para cada idioma, obtener la traducción correspondiente del plato
                    for index, language in enumerate(languages, start=1):
                        translation = DishesLang.objects.filter(dish_id=dish_id, language=language.code).first()
                        
                        if translation:
                            # Agregar la traducción al diccionario con el formato adecuado
                            translations_dict[f"TRANSLATION_{index}"] = translation.translation

                            
                allCount=0
                
                logging.info(f"articleId: {mac}, articleName: {label}, VEGETARIANO : vegetarian, ITEM_ID: {mac},  ITEM_DESCRIPTION: {itemDescription.upper()}")
                
                if exception_id != '':

                    # obtenemos el ID del plato
                    dish=Dishes.objects.filter(id=exception_id).first()

                    # Obtenemos el id del articulo a editar
                    assignment=Assignment.objects.filter(id=id).first()

                    # Obtenemos el plato original para agregar la excepcion
                    original_dish=Dishes.objects.filter(id=exception_id).first()

                    # Obtenemos las traducciones del plato
                    dishlang=DishesLang.objects.filter(dish_id=exception_id).values_list('translation', flat=True)

                    # Obtenemos el label del articulo a editar
                    labels=Label.objects.filter(id=label_id).first()


                    dish = Dishes.objects.get(id=dish_id)
                    recipes = Recipe.objects.filter(dish=dish).prefetch_related('allergens')

                    # Luego, modifica tu código actual así:
                    alergenos = alergenos_all.copy()  # Crea una copia del diccionario base

                    for recipe in recipes:
                        for allergen in recipe.allergens.all():
                            if allergen.id == 0:
                                vegetarian = "YES"
                            else:
                                allCount += 1
                                alergenos[allergen.allergen.upper()] = str(allCount)
                                vegetarian = "NO"

                    allCount=0

                    # --------------------------------------------------------------------------
                    # Consulta para obtener idiomas con estado verdadero y posición ascendente
                    # --------------------------------------------------------------------------

                    translations_dict = translation_all.copy()  # Crea una copia del diccionario base

                    # Paso 4: Obtener los idiomas activos ordenados por posición
                    languages = Language.objects.filter(status=True).order_by('position')

                    # Paso 5: Para cada idioma, obtener la traducción correspondiente del plato
                    for index, language in enumerate(languages, start=1):
                        translation = DishesLang.objects.filter(dish_id=exception_id, language=language.code).first()

                        if translation:
                            # Agregar la traducción al diccionario con el formato adecuado
                            translations_dict[f"TRANSLATION_{index}"] = translation.translation
                    
                    if dish_id == 0:
                        assigmentList.append({
                            "articleId": f"{mac}",
                            "articleName": f"{label}",
                            "data": {
                            "ITEM_DESCRIPTION": "PLATO",
                            "IS_EMPTY": "YES",
                            "HAS_EXCEPTION": "NO",
                            "VEGETARIANO" : vegetarian,
                            "EXCEPTION_NAME":"",
                            "EXCEPTION_DESCRIPTION":"",
                            "EXCEPTION_CATEGORY_03":"",
                            "ALTRAMUCES": "",
                            "APIO": "",
                            "CACAHUETES": "",
                            "CRUSTACEO": "",
                            "FRUTOS_SECOS": "",
                            "GLUTEN": "",
                            "HUEVOS": "",
                            "LACTEOS": "",
                            "MOLUSCO": "",
                            "MOSTAZA": "",
                            "PESCADO": "",
                            "SESAMO": "",
                            "SOJA": "",
                            "SULFITO": "",
                            "TRANSLATION_1" : "DISH",
                            "TRANSLATION_2" : "GERICHT",
                            "TRANSLATION_3" : "PLAT",
                            "TRANSLATION_4" : ""
                        }})
                    else:
                        assigmentList.append({
                        "articleId": f"{mac}",
                        "articleName": f"{label}",
                        "nfcUrl": f"",
                        "eans": [],
                        "data": {
                            "ITEM_ID": f"{mac}",
                            "ITEM_NAME": f"{label}",
                            "ITEM_DESCRIPTION": f"{itemDescription.upper()}",
                            "IS_EMPTY": "NO",
                            "HAS_EXCEPTION": "YES",
                            "IS_VEGETARIAN": vegetarian,
                            "EXCEPTION_NAME":f"{dish}",
                            "EXCEPTION_DESCRIPTION":", ".join(dishlang),
                            "EXCEPTION_CATEGORY_03": "",
                            **alergenos,
                            **translations_dict
                            }
                        })
                else:        
                    if dish_id == 0:
                        assigmentList.append({
                            "articleId": f"{mac}",
                            "articleName": f"{label}",
                            "data": {
                            "ITEM_DESCRIPTION": "PLATO",
                            "IS_EMPTY": "YES",
                            "HAS_EXCEPTION": "NO",
                            "VEGETARIANO" : vegetarian,
                            "EXCEPTION_NAME":"",
                            "EXCEPTION_DESCRIPTION":"",
                            "EXCEPTION_CATEGORY_03":"",
                            "ALTRAMUCES": "",
                            "APIO": "",
                            "CACAHUETES": "",
                            "CRUSTACEO": "",
                            "FRUTOS_SECOS": "",
                            "GLUTEN": "",
                            "HUEVOS": "",
                            "LACTEOS": "",
                            "MOLUSCO": "",
                            "MOSTAZA": "",
                            "PESCADO": "",
                            "SESAMO": "",
                            "SOJA": "",
                            "SULFITO": "",
                            "TRANSLATION_1" : "DISH",
                            "TRANSLATION_2" : "GERICHT",
                            "TRANSLATION_3" : "PLAT",
                            "TRANSLATION_4" : ""
                        }})
                    else:
                        assigmentList.append({
                            "articleId": f"{mac}",
                            "articleName": f"{label}",
                            "nfcUrl": f"",
                            "eans": [],
                            "data": {
                                "ITEM_ID": f"{mac}",
                                "ITEM_NAME": f"{label}",
                                "ITEM_DESCRIPTION": f"{itemDescription.upper()}",
                                "IS_EMPTY": "NO",
                                "HAS_EXCEPTION": "NO",
                                "VEGETARIANO" : vegetarian,
                                "EXCEPTION_NAME": "",
                                "EXCEPTION_DESCRIPTION": "",
                                "EXCEPTION_CATEGORY_03": "",
                                **alergenos,
                                **translations_dict
                            }
                        })
                        
        # if bulk_values:
        #     placeholders = ", ".join(["(%s, %s)"] * len(bulk_values))
        #     flat_params = []
        #     for label, new_rest in bulk_values:
        #         flat_params.extend([label, new_rest])

        #     update_sql = f"""
        #         UPDATE {schema_name}.labels_label AS l
        #         SET restaurant_id = v.new_restaurant_id
        #         FROM (
        #             VALUES
        #             {placeholders}
        #         ) AS v(label, new_restaurant_id)
        #         WHERE l.label = v.label;
        #     """
        #     with connection.cursor() as cursor:
        #         cursor.execute(update_sql, flat_params)

        eupdateUrl = f"https://eu.common.solumesl.com/common/api/v2/common/articles?company={company}&store={code}"

        eupdatePayload = json.dumps(assigmentList)

        eupdateHeaders = {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {access_token}'
        }

        eupdateResponse = requests.put(eupdateUrl, headers=eupdateHeaders, data=eupdatePayload)
        time.sleep(1)
        updateArticles=update_articles(company, code, assigmentList, access_token)
        
        messages.append(f"Company: {company}, Code: {code}, Fecha: {fecha_actual} - {hora_actual_dt}, Total de asignaciones: {total_assignments}, Excepciones para esta rueda: {total_exceptions}")
    else:
        messages.append(f"Debes estar en una franja horaria programada para cargar asignaciones o tener asignaciones validas")
        messages.append(f"Fecha: {fecha_actual}, Dia de la semana: {dia_semana}, Hora: {hora_actual_dt}")
            
        description = "Debes estar en una franja horaria programada para cargar asignaciones o tener asignaciones validas"
        Logs.objects.create(function = "labels_template", description = description, status = "Success") 
        
    return messages         

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Labels Status ------------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def labels_status(user, lbl_status, page=1, page_size=10, search=None):
    # Configuramos el esquema y obtenemos los valores de store
    set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']
    logging.info(f"Company: {company}")
    logging.info(f"Code: {code}")
    logging.info(f"Labels status: {lbl_status}")

    # Construye la URL para la solicitud
    labels_url = (f"https://eu.common.solumesl.com/common/api/v2/common/labels?company={company}&store={code}&status={lbl_status}&templateType=Discount")
    labels_payload = json.dumps({})  # Payload vacío, según tu implementación original
    labels_headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {access_token}'
    }

    # Solicitud al API
    labels_response = requests.get(labels_url, headers=labels_headers, data=labels_payload)
    if labels_response.text:
        response_json = labels_response.json()
        # Se extrae la lista de etiquetas de la respuesta
        labels = response_json.get('labelList', [])
        
        # Aplicar filtro de búsqueda por articleName si se proporcionó el parámetro search
        if search:
            search_lower = search.lower()
            filtered_labels = []
            for label in labels:
                # Se asume que cada etiqueta tiene una lista de artículos en "articleList"
                articles = label.get('articleList', [])
                # Si al menos uno de los artículos tiene un "articleName" que contiene el término de búsqueda, se incluye la etiqueta
                if any(search_lower in article.get('articleName', '').lower() for article in articles):
                    filtered_labels.append(label)
            labels = filtered_labels

        # Aplicar paginación utilizando Django Paginator
        paginator = Paginator(labels, page_size)
        page_obj = paginator.get_page(page)

        return {
            "results": list(page_obj),   # Lista de elementos de la página actual
            "total": paginator.count,      # Total de elementos después del filtrado
            "num_pages": paginator.num_pages,
            "current_page": page
        }
    else:
        return {"error": "La respuesta del API está vacía"}

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Zona Horaria -------------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def hour_zone(user):
    # 0) Esquema y store
    set_schema(user)
    store_code(user)  # define globals: company, code

    # 1) Obtener token
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return None

    # 2) Llamada al API
    url = (f"https://eu.common.solumesl.com/common/api/v2/common/store/summary?company={company}&store={code}")
    headers = {'Authorization': f"Bearer {token_data['access_token']}"}
    resp = requests.get(url, headers=headers)
    
    if resp.status_code != 200:
        logger.error(f"Error al consultar API: {resp.status_code}")
        return None

    # 3) Zona horaria
    timezone_str = resp.json()['zoneId']
    tz_store     = pytz.timezone(timezone_str)

    # 4) Ahora UTC-aware
    now_utc = datetime.now(pytz.UTC)
    # 5) Convertir a zona de la tienda
    store_dt = now_utc.astimezone(tz_store)

    # 6) Formatos de salida
    fecha_formateada = store_dt.strftime("%Y-%m-%d")
    dia_semana       = store_dt.isoweekday()  # 1=Lunes … 7=Domingo

    # 7) Reasignar hora_actual_dt con el formato deseado
    hora_actual_dt = store_dt.time()

    return timezone_str, fecha_formateada, dia_semana, hora_actual_dt

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def serializers_hour_zone(user):
    # 0) Esquema y store
    set_schema(user)
    store_code(user)  # define globals: company, code

    # 1) Obtener token
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return None

    # 2) Llamada al API
    url = (
        f"https://eu.common.solumesl.com/common/api/v2/common/"
        f"store/summary?company={company}&store={code}"
    )
    headers = {'Authorization': f"Bearer {token_data['access_token']}"}
    resp = requests.get(url, headers=headers)
    if resp.status_code != 200:
        logger.error(f"Error al consultar API: {resp.status_code}")
        return None

    # 3) Zona horaria
    timezone_str = resp.json()['zoneId']
    tz_store     = pytz.timezone(timezone_str)

    # 4) Ahora UTC-aware
    now_utc = datetime.now(pytz.UTC)
    # 5) Convertir a zona de la tienda
    store_dt = now_utc.astimezone(tz_store)

    # 6) Formatos de salida
    fecha_formateada = store_dt.strftime("%Y-%m-%d")
    dia_semana       = store_dt.isoweekday()  # 1=Lunes … 7=Domingo

    # 7) Reasignar hora_actual_dt con el formato deseado
    hora_actual_dt = store_dt.time()

    return timezone_str, fecha_formateada, dia_semana, hora_actual_dt

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Clear Dish ---------------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def restore_dish(day_of_week, articleId, dish_id, time_id, user):

    # Configuramos el esquema y obtenemos los valores de store
    set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']

    allCount = 0

    recived_dish = dish_id
    
    messages=[]
            
    logging.info(f"day_of_week: {day_of_week}")
    logging.info(f"articleId: {articleId}")
    logging.info(f"dish_id: {dish_id}")
    logging.info(f"time_id: {time_id}")
    
    l = Label.objects.filter(mac=articleId).first()
    
    assignment=Assignment.objects.filter(day_of_week=day_of_week, label_id=l.id, time_id=time_id).first()
    
    logging.info(f"Query de la tabla asignaciones obtenisda:")
    logging.info(f"day_of_week: {assignment.day_of_week}")
    logging.info(f"label_id: {assignment.label_id}")
    logging.info(f"dish_id: {assignment.dish_id}")
    logging.info(f"time_id: {assignment.time_id}")
    
    # Buscar la zona horaria
    timezone, fecha_actual, dia_semana, hora_actual_dt = hour_zone(user)
    time.sleep(2)
    
    # Verificar si existe una semana que contenga la fecha actual            
    existe_fecha = Week.objects.filter(start_date__lte=fecha_actual,end_date__gte=fecha_actual,active=True).annotate(rango_dias=F('end_date') - F('start_date')).filter(rango_dias__lt=timedelta(days=365))
    valid_weeks = Week.objects.filter(start_date__lte=fecha_actual,end_date__gte=fecha_actual,active=True).annotate(rango_dias=F('end_date') - F('start_date')).filter(rango_dias__gte=timedelta(days=365))
            
    if existe_fecha:
        week_ids = existe_fecha.values_list('id', flat=True)
        messages.append(f"El weekID es: {week_ids}")
                
        a = Assignment.objects.filter(day_of_week=dia_semana, label_id=l.id, dish_id=assignment.dish_id, time_id=time_id, week_id__in=week_ids).first()
                
    else:
            
        week_ids = valid_weeks.values_list('id', flat=True)
        messages.append(f"El weekID es: {week_ids}")
                    
        a = Assignment.objects.filter(day_of_week=day_of_week, label_id=l.id, dish_id=assignment.dish_id, time_id=time_id).first()
        
    d = Dishes.objects.filter(id=a.dish_id).first()
    
    tn=Time.objects.filter(id=time_id).first()
    
    logging.info(f"Time obtenido de la consulta Time: {tn.id}")
    logging.info(f"Dia de la semana actual: {dia_semana}")
    logging.info(f"Dia de la semana de la asignacion enviada: {day_of_week}")
    
    t = Time.objects.filter(start__lte=hora_actual_dt, end__gte=hora_actual_dt).first()
    
    if t is not None:
        # Se encontró un objeto
        logging.info(f"Time obtenido de la consulta Time en el rango de horas activo: {t.name}")
        logging.info(f"ID obtenido de la consulta Time en el rango de horas activo: {t.id}")
        logging.info(f"t.id es igual a: {t.name}")
        logging.info(f"t.id es igual a: {tn.name}")
        
        if dia_semana == a.day_of_week and t.id == tn.id:
            
            logging.info("La data se cargara en SINCRONI y SOLUM")


            if recived_dish != 0:
                Exception.objects.filter(assignment_id=a.id).delete()

            dish = Dishes.objects.get(id=a.dish_id)

            recipes = Recipe.objects.filter(dish=dish).prefetch_related('allergens')

            alergenos_all = {
                "ALTRAMUCES": "",
                "APIO": "",
                "CACAHUETES": "",
                "CRUSTACEO": "",
                "FRUTOS_SECOS": "",
                "GLUTEN": "",
                "HUEVOS": "",
                "LACTEOS": "",
                "MOLUSCO": "",
                "MOSTAZA": "",
                "PESCADO": "",
                "SESAMO": "",
                "SOJA": "",
                "SULFITO": ""
            }

            translation_all = {
                "TRANSLATION_1": "",
                "TRANSLATION_2": "",
                "TRANSLATION_3": "",
                "TRANSLATION_4": ""
            }

            allergens_dict = alergenos_all.copy()

            for recipe in recipes:
                for allergen in recipe.allergens.all():
                    allCount += 1
                    allergens_dict[allergen.allergen.upper()] = str(allCount)


            # --------------------------------------------------------------------------
            # Consulta para obtener idiomas con estado verdadero y posición ascendente
            # --------------------------------------------------------------------------

            translations_dict = translation_all.copy()

            dish_id = a.dish_id

            # Paso 4: Obtener los idiomas activos ordenados por posición
            languages = Language.objects.filter(status=True).order_by('position')

            # Paso 5: Para cada idioma, obtener la traducción correspondiente del plato
            for index, language in enumerate(languages, start=1):
                translation = DishesLang.objects.filter(dish_id=dish_id, language=language.code).first()

                if translation:
                    # Agregar la traducción al diccionario con el formato adecuado
                    translations_dict[f"TRANSLATION_{index}"] = translation.translation

            clearUrl = f"https://eu.common.solumesl.com/common/api/v2/common/articles?company={company}&store={code}"

            if dish_id == 0:
                clearPayload = json.dumps([{
                    "articleId": f"{articleId}",
                    "data": {
                    "ITEM_DESCRIPTION": "PLATO",
                    "IS_EMPTY": "YES",
                    "HAS_EXCEPTION": "NO",
                    "EXCEPTION_NAME":"",
                    "EXCEPTION_DESCRIPTION":"",
                    "EXCEPTION_CATEGORY_03":"",
                    "ALTRAMUCES": "",
                    "APIO": "",
                    "CACAHUETES": "",
                    "CRUSTACEO": "",
                    "FRUTOS_SECOS": "",
                    "GLUTEN": "",
                    "HUEVOS": "",
                    "LACTEOS": "",
                    "MOLUSCO": "",
                    "MOSTAZA": "",
                    "PESCADO": "",
                    "SESAMO": "",
                    "SOJA": "",
                    "SULFITO": "",
                    "TRANSLATION_1" : "DISH",
                    "TRANSLATION_2" : "GERICHT",
                    "TRANSLATION_3" : "PLAT",
                    "TRANSLATION_4" : ""
                }}])
            else:
                clearPayload = json.dumps([{
                    "articleId": f"{articleId}",
                    "data": {
                    "ITEM_DESCRIPTION": f"{d.dish.upper()}",
                    "IS_EMPTY": "NO",
                    "HAS_EXCEPTION": "NO",
                    "EXCEPTION_NAME":"",
                    "EXCEPTION_DESCRIPTION":"",
                    "EXCEPTION_CATEGORY_03":"",
                    **allergens_dict,
                    **translations_dict
                }}])

            clearHeaders = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {access_token}'
            }

            clearResponse = requests.put(clearUrl, headers=clearHeaders, data=clearPayload)

            if clearResponse.status_code==200:

                if recived_dish == 0:

                    Assignment.objects.filter(day_of_week=day_of_week, label_id=l.id, dish_id=assignment.dish_id, time_id=time_id).update(dish_id=0)

                    Logs.objects.create(function = "restore_dish", description = f"La asignacion {l.label} se ha restaurado con el dish ID 0", status = "Success")

                    messages.append(f"Se ha limpiado el articleId {articleId} - {l.label} con el dish ID 0 (Plato)")
                    
                else:
                    Logs.objects.create(function = "restore_dish", description = f"Se ha restaurado el articleId {articleId} - {l.label} al valor por defecto de la asignacion {d.dish}", status = "Success")

                    messages.append(f"Se ha restaurado el articleId {articleId} - {l.label} al valor por defecto de la asignacion {d.dish}")

            else:

                Logs.objects.create(function = "restore_dish", description = f"No ha restaurado el articleId {articleId}", status = "Error")

                messages.append(f"No ha restaurado el articleId {articleId}")

        else:
            if recived_dish != 0:
                Exception.objects.filter(assignment_id=a.id).delete()
                
            if recived_dish == 0:
            
                logging.info(f"Assignment ID: {a.id}, Time ID: {time_id}") 

                Assignment.objects.filter(day_of_week=day_of_week, label_id=l.id, dish_id=assignment.dish_id, time_id=time_id).update(dish_id=0)

                Logs.objects.create(function = "restore_dish", description = f"La asignacion {l.label} se ha restaurado con el dish ID 0", status = "Success")

                messages.append(f"Se ha limpiado el articleId {articleId} - {l.label} con el dish ID 0 (Plato)")
                    
            else:
                logging.info(f"Assignment ID: {a.id}, Time ID: {time_id}") 
            
                Logs.objects.create(function = "restore_dish", description = f"Se ha restaurado el articleId {articleId} - {l.label} al valor por defecto de la asignacion {d.dish}", status = "Success")

                messages.append(f"Se ha restaurado el articleId {articleId} - {l.label} al valor por defecto de la asignacion {d.dish}")

            # No se encontró ningún objeto
            logging.info("No se cargara la excepcion a solum solo se cargara en sincroni")
                

    return messages

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Clear Dish Assignment ----------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def restore_dish_assignment(day_of_week, articleId, dish_id, time_id, user):

    # Configuramos el esquema y obtenemos los valores de store
    set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    allCount = 0

    recived_dish = dish_id
    
    messages=[]
            
    logging.info(f"day_of_week: {day_of_week}")
    logging.info(f"articleId: {articleId}")
    logging.info(f"dish_id: {dish_id}")
    logging.info(f"time_id: {time_id}")
    
    l = Label.objects.filter(mac=articleId).first()
    
    assignment=Assignment.objects.filter(day_of_week=day_of_week, label_id=l.id, time_id=time_id).first()
    
    logging.info(f"Query de la tabla asignaciones obtenisda:")
    logging.info(f"day_of_week: {assignment.day_of_week}")
    logging.info(f"label_id: {assignment.label_id}")
    logging.info(f"dish_id: {assignment.dish_id}")
    logging.info(f"time_id: {assignment.time_id}")
    
    # Buscar la zona horaria
    timezone, fecha_actual, dia_semana, hora_actual_dt = hour_zone(user)
    time.sleep(2)
    
    # Verificar si existe una semana que contenga la fecha actual            
    existe_fecha = Week.objects.filter(start_date__lte=fecha_actual,end_date__gte=fecha_actual,active=True).annotate(rango_dias=F('end_date') - F('start_date')).filter(rango_dias__lt=datetime.timedelta(days=365))
    valid_weeks = Week.objects.filter(start_date__lte=fecha_actual,end_date__gte=fecha_actual,active=True).annotate(rango_dias=F('end_date') - F('start_date')).filter(rango_dias__gte=timedelta(days=365))
            
    if existe_fecha:
        week_ids = existe_fecha.values_list('id', flat=True)
        messages.append(f"El weekID es: {week_ids}")
                
        a = Assignment.objects.filter(day_of_week=dia_semana, label_id=l.id, dish_id=assignment.dish_id, time_id=time_id, week_id__in=week_ids).first()
                
    else:
            
        week_ids = valid_weeks.values_list('id', flat=True)
        messages.append(f"El weekID es: {week_ids}")
                    
        a = Assignment.objects.filter(day_of_week=day_of_week, label_id=l.id, dish_id=assignment.dish_id, time_id=time_id).first()
        
    d = Dishes.objects.filter(id=a.dish_id).first()
    
    tn=Time.objects.filter(id=time_id).first()
    
    logging.info(f"Time obtenido de la consulta Time: {tn.id}")
    logging.info(f"Dia de la semana actual: {dia_semana}")
    logging.info(f"Dia de la semana de la asignacion enviada: {day_of_week}")
    
    t = Time.objects.filter(start__lte=hora_actual_dt, end__gte=hora_actual_dt).first()
    
    if t is not None:
        # Se encontró un objeto
        logging.info(f"Time obtenido de la consulta Time en el rango de horas activo: {t.name}")
        logging.info(f"ID obtenido de la consulta Time en el rango de horas activo: {t.id}")
        logging.info(f"t.id es igual a: {t.name}")
        logging.info(f"t.id es igual a: {tn.name}")
        
            
        logging.info("La data se cargara en SINCRONI y SOLUM")


        if recived_dish != 0:
            Exception.objects.filter(assignment_id=a.id).delete()

        
        if recived_dish == 0:

            Assignment.objects.filter(day_of_week=day_of_week, label_id=l.id, dish_id=assignment.dish_id, time_id=time_id).update(dish_id=0)

            Logs.objects.create(function = "restore_dish_assignment", description = f"La asignacion {l.label} se ha restaurado con el dish ID 0", status = "Success")

            messages.append(f"Se ha limpiado el articleId {articleId} - {l.label} con el dish ID 0 (Plato)")
            
        else:
            Logs.objects.create(function = "restore_dish", description = f"Se ha restaurado el articleId {articleId} - {l.label} al valor por defecto de la asignacion {d.dish}", status = "Success")

            messages.append(f"Se ha restaurado el articleId {articleId} - {l.label} al valor por defecto de la asignacion {d.dish}")
                

    return messages

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def delete_exception(day_of_week, articleId, dish_id, time_id, user):

    # Configuramos el esquema y obtenemos los valores de store
    set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']

    allCount = 0
    recived_dish = dish_id
    messages=[]
        
    logging.info(f"day_of_week: {day_of_week}")
    logging.info(f"articleId: {articleId}")
    logging.info(f"dish_id: {dish_id}")
    logging.info(f"time_id: {time_id}")

    l = Label.objects.filter(mac=articleId).first()
    
    messages.append(f"Day of week: {day_of_week}")           
    messages.append(f"Label MAC: {l.mac}")
    messages.append(f"Dish ID: {dish_id}")
    messages.append(f"Time ID: {time_id}")
    
    assignment=Assignment.objects.filter(day_of_week=day_of_week, label_id=l.id, time_id=time_id).first()
    
    # Buscar la zona horaria
    timezone, fecha_actual, dia_semana, hora_actual_dt = hour_zone(user)
    time.sleep(2)

    
    # Verificar si existe una semana que contenga la fecha actual            
    existe_fecha = Week.objects.filter(start_date__lte=fecha_actual,end_date__gte=fecha_actual,active=True).annotate(rango_dias=F('end_date') - F('start_date')).filter(rango_dias__lt=timedelta(days=365))
    valid_weeks = Week.objects.filter(start_date__lte=fecha_actual,end_date__gte=fecha_actual, active=True).annotate(rango_dias=F('end_date') - F('start_date')).filter(rango_dias__gte=timedelta(days=365))
            
    if existe_fecha:
        week_ids = existe_fecha.values_list('id', flat=True)
        messages.append(f"El weekID es: {week_ids}")
                
        a = Assignment.objects.filter(day_of_week=dia_semana, label_id=l.id, dish_id=assignment.dish_id, time_id=time_id, week_id__in=week_ids).first()
        logging.info(f"Assignment ID: {a.id}")
        Exception.objects.filter(assignment_id=a.id).delete()
                
    else:
    
        week_ids = valid_weeks.values_list('id', flat=True)
        messages.append(f"El weekID es: {week_ids}")
                    
        a = Assignment.objects.filter(day_of_week=day_of_week, label_id=l.id, dish_id=assignment.dish_id, time_id=time_id, week_id__in=week_ids).first()
        messages.append(f"Assignment ID: {assignment.id}")
        messages.append(f"Assignment ID de la segunda consulta: {a.id}")
        messages.append(f'Se eliminara la excepcion con la asignacion nro: {a.id}')
        
        instance = Exception.objects.filter(assignment_id=assignment.id).first()

        # Verificar si se encontró una instancia
        if instance:
            messages.append(f'Assignment ID a eliminar es: {instance.id}')
            instance.delete()  # Eliminar la instancia
            messages.append('Se eliminó la instancia de Exception.')
        else:
            messages.append('No se encontró ninguna instancia de Exception para eliminar.')
        
        
    d = Dishes.objects.filter(id=a.dish_id).first()
    
    tn=Time.objects.filter(id=time_id).first()
    
    t = Time.objects.filter(start__lte=hora_actual_dt, end__gte=hora_actual_dt).first()
    
    if t is not None:
        
        if dia_semana == a.day_of_week and t.id == tn.id:
            
            logging.info("La data se cargara en SINCRONI y SOLUM")


            if recived_dish != 0:
                Exception.objects.filter(assignment_id=a.id).delete()

            dish = Dishes.objects.get(id=a.dish_id)

            recipes = Recipe.objects.filter(dish=dish).prefetch_related('allergens')

            alergenos_all = {
                "ALTRAMUCES": "",
                "APIO": "",
                "CACAHUETES": "",
                "CRUSTACEO": "",
                "FRUTOS_SECOS": "",
                "GLUTEN": "",
                "HUEVOS": "",
                "LACTEOS": "",
                "MOLUSCO": "",
                "MOSTAZA": "",
                "PESCADO": "",
                "SESAMO": "",
                "SOJA": "",
                "SULFITO": ""
            }

            translation_all = {
                "TRANSLATION_1": "",
                "TRANSLATION_2": "",
                "TRANSLATION_3": "",
                "TRANSLATION_4": ""
            }

            allergens_dict = alergenos_all.copy()

            for recipe in recipes:
                for allergen in recipe.allergens.all():
                    allCount += 1
                    allergens_dict[allergen.allergen.upper()] = str(allCount)


            # --------------------------------------------------------------------------
            # Consulta para obtener idiomas con estado verdadero y posición ascendente
            # --------------------------------------------------------------------------

            translations_dict = translation_all.copy()

            dish_id = a.dish_id

            # Paso 4: Obtener los idiomas activos ordenados por posición
            languages = Language.objects.filter(status=True).order_by('position')

            # Paso 5: Para cada idioma, obtener la traducción correspondiente del plato
            for index, language in enumerate(languages, start=1):
                translation = DishesLang.objects.filter(dish_id=dish_id, language=language.code).first()

                if translation:
                    # Agregar la traducción al diccionario con el formato adecuado
                    translations_dict[f"TRANSLATION_{index}"] = translation.translation

            clearUrl = f"https://eu.common.solumesl.com/common/api/v2/common/articles?company={company}&store={code}"

            if a.dish_id == 0:
                clearPayload = json.dumps([{
                    "articleId": f"{articleId}",
                    "data": {
                    "ITEM_DESCRIPTION": "PLATO",
                    "IS_EMPTY": "YES",
                    "HAS_EXCEPTION": "NO",
                    "EXCEPTION_NAME":"",
                    "EXCEPTION_DESCRIPTION":"",
                    "EXCEPTION_CATEGORY_03":"",
                    "ALTRAMUCES": "",
                    "APIO": "",
                    "CACAHUETES": "",
                    "CRUSTACEO": "",
                    "FRUTOS_SECOS": "",
                    "GLUTEN": "",
                    "HUEVOS": "",
                    "LACTEOS": "",
                    "MOLUSCO": "",
                    "MOSTAZA": "",
                    "PESCADO": "",
                    "SESAMO": "",
                    "SOJA": "",
                    "SULFITO": "",
                    "TRANSLATION_1" : "DISH",
                    "TRANSLATION_2" : "GERICHT",
                    "TRANSLATION_3" : "PLAT",
                    "TRANSLATION_4" : ""
                }}])
            else:
                clearPayload = json.dumps([{
                    "articleId": f"{articleId}",
                    "data": {
                    "ITEM_DESCRIPTION": f"{d.dish.upper()}",
                    "IS_EMPTY": "NO",
                    "HAS_EXCEPTION": "NO",
                    "EXCEPTION_NAME":"",
                    "EXCEPTION_DESCRIPTION":"",
                    "EXCEPTION_CATEGORY_03":"",
                    **allergens_dict,
                    **translations_dict
                }}])

            clearHeaders = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {access_token}'
            }

            clearResponse = requests.put(clearUrl, headers=clearHeaders, data=clearPayload)

            if clearResponse.status_code==200:
            
                Logs.objects.create(function = "restore_dish", description = f"Se ha restaurado el articleId {articleId} - {l.label} al valor por defecto de la asignacion {d.dish}", status = "Success")
                messages.append(f"Se ha restaurado el articleId {articleId} - {l.label} al valor por defecto de la asignacion {d.dish}")

            else:

                Logs.objects.create(function = "restore_dish", description = f"No ha restaurado el articleId {articleId}", status = "Error")
                messages.append(f"No ha restaurado el articleId {articleId}")

        else:
            Logs.objects.create(function = "restore_dish", description = f"Se ha restaurado el articleId {articleId} - {l.label} al valor por defecto de la asignacion {d.dish}", status = "Success")
            messages.append(f"Se ha restaurado el articleId {articleId} - {l.label} al valor por defecto de la asignacion {d.dish}")

            # No se encontró ningún objeto
            logging.info("No se cargara la excepcion a solum solo se cargara en sincroni")
                
    return messages

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Clear Dish ---------------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def clear_dish(articleId, user):

    # Configuramos el esquema y obtenemos los valores de store
    set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']

    dish_id = articleId

    print(dish_id)

    clearUrl = f"https://eu.common.solumesl.com/common/api/v2/common/articles?company={company}&store={code}"

    clearPayload = json.dumps([{
        "articleId": f"{dish_id}",
        "data": {
        "ITEM_DESCRIPTION": "",
        "IS_EMPTY": "YES",
        "HAS_EXCEPTION": "NO",
        "EXCEPTION_NAME":"",
        "EXCEPTION_DESCRIPTION":"",
        "EXCEPTION_CATEGORY_03":"",
        "ALTRAMUCES": "",
        "APIO": "",
        "CACAHUETES": "",
        "CRUSTACEO": "",
        "FRUTOS_SECOS": "",
        "GLUTEN": "",
        "HUEVOS": "",
        "LACTEOS": "",
        "MOLUSCO": "",
        "MOSTAZA": "",
        "PESCADO": "",
        "SESAMO": "",
        "SOJA": "",
        "SULFITO": "",
        "TRANSLATION_1" : "",
        "TRANSLATION_2" : "",
        "TRANSLATION_3" : "",
        "TRANSLATION_4" : ""
    }}])

    clearHeaders = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {access_token}'
    }

    clearResponse = requests.put(clearUrl, headers=clearHeaders, data=clearPayload)

    if clearResponse.status_code==200:

        messages=f"Se ha restaurado el articleId {dish_id}"

    else:
        messages=f"No ha restaurado el articleId {dish_id}"

    print(clearResponse.status_code)

    return messages

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Clear Exception ----------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def clear_exception(company, code, access_token, user):

    set_schema(user)

    print(f"Function clear_exception")
    print(f"Company: {company}")
    print(f"Code: {code}")

    messages=[]

    updateList=[]

    response=None

    # consultar los labels asignados para ser limpiados
    assignmentsLabelsUrl=f"https://eu.common.solumesl.com/common/api/v2/common/labels/assigned?company={company}&store={code}&size=500"
    assignmentsLabelspayload = json.dumps({})
    assignmentsLabelsHeaders={
        'Authorization': f'Bearer {access_token}'
    }
    assignmentsLabelsResponse = requests.get(assignmentsLabelsUrl, headers=assignmentsLabelsHeaders, data=assignmentsLabelspayload)

    response=assignmentsLabelsResponse.status_code

    if assignmentsLabelsResponse.status_code == 200:
        response_json = assignmentsLabelsResponse.json()
        assigned_articles = []

        if response_json:

            for label in response_json["assignedLabelsList"]:

                assigned_articles.extend(label["assignedArticles"])

                labelId = label["assignedArticles"]

                # print(labelId[0])

                updateList.append({
                    "articleId": f"{labelId[0]}",
                    "data": {
                    "IS_EMPTY": "YES",
                    "HAS_EXCEPTION": "NO",
                    "EXCEPTION_NAME":"",
                    "EXCEPTION_DESCRIPTION":"",
                    "EXCEPTION_CATEGORY_03":"",
                    "ALTRAMUCES": "",
                    "APIO": "",
                    "CACAHUETES": "",
                    "CRUSTACEO": "",
                    "FRUTOS_SECOS": "",
                    "GLUTEN": "",
                    "HUEVOS": "",
                    "LACTEOS": "",
                    "MOLUSCO": "",
                    "MOSTAZA": "",
                    "PESCADO": "",
                    "SESAMO": "",
                    "SOJA": "",
                    "SULFITO": "",
                    }
                })

            updateUrlx = f"https://eu.common.solumesl.com/common/api/v2/common/articles?company={company}&store={code}"

            updatePayloadx = json.dumps(updateList)

            updateHeadersx = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {access_token}'
            }

            updateResponsex = requests.put(updateUrlx, headers=updateHeadersx, data=updatePayloadx)

            response=updateResponsex.status_code

    return response

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Actualizar Label_1 -------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def update_articles(company, code, assigmentList, access_token):

    xurl = f"https://eu.common.solumesl.com/common/api/v2/common/articles?company={company}&store={code}"

    xpayload = json.dumps(assigmentList)

    xheaders = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {access_token}'
    }

    xresponse = requests.put(xurl, headers=xheaders, data=xpayload)

    return xresponse.status_code

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Exception ----------------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Exception ----------------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------

def add_exception(dish_id, label_id, assignment_id, time_name, user):

    # Configuramos el esquema y obtenemos los valores de store
    set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']

    allCount=0
    
    vegetarian = "NO"

    updateList=[]
    
    assigmentList = []
    
    messages=[]

    current_date = date.today()
            
    vegetarian = "NO"

    timezone, fecha_actual, dia_semana, hora_actual_dt = hour_zone(user)
    time.sleep(1)
    
    logging.info(f"dish_id: {dish_id}")
    logging.info(f"label_id: {label_id}")
    logging.info(f"assignment_id: {assignment_id}")
    logging.info(f"time: {time_name}")
    
    # dish_id: 161
    # label_id: 20
    # assignment_id: 560
    # time: Almuerzo
    
    tn=Time.objects.filter(name=time_name).first()
    logging.info(f"Time obtenido de la consulta Time: {tn.id}")
    
    a=Assignment.objects.filter(id=assignment_id).first()
    
    logging.info(f"Dia de la semana actual: {dia_semana}")
    
    logging.info(f"Dia de la semana de la asignacion enviada: {a.day_of_week}")
    
    t = Time.objects.filter(start__lte=hora_actual_dt, end__gte=hora_actual_dt).first()
    if t is not None:
        # Se encontró un objeto
        logging.info(f"Time obtenido de la consulta Time en el rango de horas activo: {t.name}")
        logging.info(f"ID obtenido de la consulta Time en el rango de horas activo: {t.id}")
        
        if dia_semana == a.day_of_week and t.id == tn.id:
            alergenos_all = {
                "ALTRAMUCES": "",
                "APIO": "",
                "CACAHUETES": "",
                "CRUSTACEO": "",
                "FRUTOS_SECOS": "",
                "GLUTEN": "",
                "HUEVOS": "",
                "LACTEOS": "",
                "MOLUSCO": "",
                "MOSTAZA": "",
                "PESCADO": "",
                "SESAMO": "",
                "SOJA": "",
                "SULFITO": ""
            }

            translation_all = {
                "TRANSLATION_1": "",
                "TRANSLATION_2": "",
                "TRANSLATION_3": "",
                "TRANSLATION_4": ""
            }

            # obtenemos el ID del plato
            dish=Dishes.objects.filter(id=dish_id).first()

            # Obtenemos el id del articulo a editar
            assignment=Assignment.objects.filter(id=assignment_id).first()

            # Obtenemos el plato original para agregar la excepcion
            original_dish=Dishes.objects.filter(id=assignment.dish_id).first()

            # Obtenemos las traducciones del plato
            dishlang=DishesLang.objects.filter(dish_id=dish_id).values_list('translation', flat=True)

            # Obtenemos el label del articulo a editar
            labels=Label.objects.filter(id=label_id).first()
            
            messages.append(f"Template: {labels.template.name}")

            dish = Dishes.objects.get(id=dish_id)
                
            recipes = Recipe.objects.filter(dish=dish).prefetch_related('allergens')

            alergenos = alergenos_all.copy()  # Crea una copia del diccionario base

            for recipe in recipes:
                for allergen in recipe.allergens.all():
                    if allergen.id == 0:
                        vegetarian = "YES"
                        messages.append(f"Is vegetarian: {vegetarian}")
                    else:
                        allCount += 1
                        alergenos[allergen.allergen.upper()] = str(allCount)
                        vegetarian = "NO"
                        messages.append(f"Is vegetarian: {vegetarian}")

            allCount=0

            # --------------------------------------------------------------------------
            # Consulta para obtener idiomas con estado verdadero y posición ascendente
            # --------------------------------------------------------------------------
            translations_dict = translation_all.copy()

            # Paso 4: Obtener los idiomas activos ordenados por posición
            languages = Language.objects.filter(status=True).order_by('position')

            # Paso 5: Para cada idioma, obtener la traducción correspondiente del plato
            for index, language in enumerate(languages, start=1):
                translation = DishesLang.objects.filter(dish_id=dish_id, language=language.code).first()

                if translation:
                    # Agregar la traducción al diccionario con el formato adecuado
                    translations_dict[f"TRANSLATION_{index}"] = translation.translation


            exceptionUrl = f"https://eu.common.solumesl.com/common/api/v2/common/articles?company={company}&store={code}"

            if dish_id == 0:
                exceptionPayload = json.dumps([{
                    "articleId": f"{labels.mac}",
                    "data": {
                        "IS_EMPTY": "YES",
                        "HAS_EXCEPTION": "YES",
                        "VEGETARIANO" : vegetarian,
                        "EXCEPTION_NAME":f"",
                        "EXCEPTION_DESCRIPTION":", ".join(dishlang),
                        **alergenos,
                        **translations_dict
                    }
                }])
            else:
                exceptionPayload = json.dumps([{
                    "articleId": f"{labels.mac}",
                    "data": {
                        "IS_EMPTY": "NO",
                        "HAS_EXCEPTION": "YES",
                        "IS_VEGETARIAN": vegetarian,
                        "EXCEPTION_NAME":f"{dish.dish}",
                        "EXCEPTION_DESCRIPTION":", ".join(dishlang),
                        **alergenos,
                        **translations_dict
                    }
                }])
                
            assigmentList.append({"articleIdList": [f"{labels.mac}"], "labelCode": f"{labels.mac}", "templateName": labels.template.name})

            exceptionHeaders = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {access_token}'
            }

            exceptionResponse = requests.put(exceptionUrl, headers=exceptionHeaders, data=exceptionPayload)
            
            logging.info(f"code: {exceptionResponse.status_code}")

            if exceptionResponse.status_code == 200:
                
                time.sleep(2)
                
                createAssignments, response_message = create_assignments(company, code, assigmentList, access_token, user)
                
                if createAssignments == 200:
                    messages.append(f"Se han creado las asignaciones correctamente")
                    
                DayliMenu.objects.filter(date=current_date, turn=time_name).update(exception=F('exception') + 1)
                    
                messages.append(f"Plato cargado correctamente a SOLUM Plato ID: {dish_id}, Label: {labels.mac}/{labels.label}")
    else:
        # No se encontró ningún objeto
        logging.info("No se cargara la excepcion a solum solo se cargara en sincroni")
                
    return messages

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def edit_exception(dish_id, label_id, assignment_id, time_name, user):

    set_schema(user)
    store_code(user)  
    
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    access_token = token_data['access_token']

    allCount=0
    
    vegetarian = "NO"

    updateList=[]
    
    messages=[]

    current_date=datetime.date.today()
            
    timezone, fecha_actual, dia_semana, hora_actual_dt = hour_zone(user)
    time.sleep(1)
    
    tn=Time.objects.filter(name=time_name).first()
    
    a=Assignment.objects.filter(id=assignment_id).first()
    
    t = Time.objects.filter(start__lte=hora_actual_dt, end__gte=hora_actual_dt).first()
    
    logging.info(f"id: {id}")
    logging.info(f"dish_id: {dish_id}")
    logging.info(f"label_id: {label_id}")
    logging.info(f"assignment_id: {assignment_id}")
    logging.info(f"time: {time_name}")
    logging.info(f"Time obtenido de la consulta Time: {tn.id}")
    logging.info(f"Dia de la semana actual: {dia_semana}")
    logging.info(f"Dia de la semana de la asignacion enviada: {a.day_of_week}")
    
    if t is not None:
        # Se encontró un objeto
        logging.info(f"Time obtenido de la consulta Time en el rango de horas activo: {t.name}")
        logging.info(f"ID obtenido de la consulta Time en el rango de horas activo: {t.id}")
        
        if dia_semana == a.day_of_week and t.id == tn.id:
            
            logging.info("La data se cargara en SINCRONI y SOLUM")
            
            alergenos_all = {
                "ALTRAMUCES": "",
                "APIO": "",
                "CACAHUETES": "",
                "CRUSTACEO": "",
                "FRUTOS_SECOS": "",
                "GLUTEN": "",
                "HUEVOS": "",
                "LACTEOS": "",
                "MOLUSCO": "",
                "MOSTAZA": "",
                "PESCADO": "",
                "SESAMO": "",
                "SOJA": "",
                "SULFITO": ""
            }

            translation_all = {
                "TRANSLATION_1": "",
                "TRANSLATION_2": "",
                "TRANSLATION_3": "",
                "TRANSLATION_4": ""
            }

            # obtenemos el ID del plato
            dish=Dishes.objects.filter(id=dish_id).first()

            # Obtenemos el id del articulo a editar
            assignment=Assignment.objects.filter(id=assignment_id).first()

            # Obtenemos el plato original para agregar la excepcion
            original_dish=Dishes.objects.filter(id=assignment.dish_id).first()

            # Obtenemos las traducciones del plato
            dishlang=DishesLang.objects.filter(dish_id=dish_id).values_list('translation', flat=True)

            # Obtenemos el label del articulo a editar
            labels=Label.objects.filter(id=label_id).first()

            dish = Dishes.objects.get(id=dish_id)
                
            recipes = Recipe.objects.filter(dish=dish).prefetch_related('allergens')

            alergenos = alergenos_all.copy()  # Crea una copia del diccionario base

            for recipe in recipes:
                for allergen in recipe.allergens.all():
                    if allergen.id == 0:
                        vegetarian = "YES"
                        messages.append(f"Is vegetarian: {vegetarian}")
                    else:
                        allCount += 1
                        alergenos[allergen.allergen.upper()] = str(allCount)
                        vegetarian = "NO"
                        messages.append(f"Is vegetarian: {vegetarian}")

            allCount=0

            # --------------------------------------------------------------------------
            # Consulta para obtener idiomas con estado verdadero y posición ascendente
            # --------------------------------------------------------------------------
            translations_dict = translation_all.copy()

            # Paso 4: Obtener los idiomas activos ordenados por posición
            languages = Language.objects.filter(status=True).order_by('position')

            # Paso 5: Para cada idioma, obtener la traducción correspondiente del plato
            for index, language in enumerate(languages, start=1):
                translation = DishesLang.objects.filter(dish_id=dish_id, language=language.code).first()

                if translation:
                    # Agregar la traducción al diccionario con el formato adecuado
                    translations_dict[f"TRANSLATION_{index}"] = translation.translation


            exceptionUrl = f"https://eu.common.solumesl.com/common/api/v2/common/articles?company={company}&store={code}"

            if dish_id == 0:
                exceptionPayload = json.dumps([{
                    "articleId": f"{labels.mac}",
                    "data": {
                        "IS_EMPTY": "YES",
                        "HAS_EXCEPTION": "YES",
                        "VEGETARIANO" : vegetarian,
                        "EXCEPTION_NAME":f"",
                        "EXCEPTION_DESCRIPTION":", ".join(dishlang),
                        **alergenos,
                        **translations_dict
                    }
                }])
            else:
                exceptionPayload = json.dumps([{
                    "articleId": f"{labels.mac}",
                    "data": {
                        "IS_EMPTY": "NO",
                        "HAS_EXCEPTION": "YES",
                        "IS_VEGETARIAN": vegetarian,
                        "EXCEPTION_NAME":f"{dish.dish}",
                        "EXCEPTION_DESCRIPTION":", ".join(dishlang),
                        **alergenos,
                        **translations_dict
                    }
                }])

            exceptionHeaders = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {access_token}'
            }

            exceptionResponse = requests.put(exceptionUrl, headers=exceptionHeaders, data=exceptionPayload)

            if exceptionResponse == 200:
                        
                messages.append(f"Plato cargado correctamente a SOLUM Plato ID: {dish_id}, Label: {labels.mac}/{labels.label}")
    
        else:
            logging.info("Se cargara la excepcion a SINCRONI estos cambios surtiran efecto en SOLUM cuando corra su rueda respectiva")
    
    else:
        # No se encontró ningún objeto
        logging.info("No se cargara la excepcion a solum solo se cargara en sincroni")
                
    return messages

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def read_exception(access_token, user):

    store_code(user)

    updateList=[]
    exceptionList=[]

    allCount=0
    
    vegetarian = "NO"

    timezone, fecha_actual, dia_semana, hora_actual_dt = hour_zone(user)
    
    # — Normalizar hora_actual_dt para que sea siempre datetime.time —
    if isinstance(hora_actual_dt, str):
        # intenta ISO datetime "YYYY-MM-DDTHH:MM:SS[.ffffff]" o "HH:MM:SS" 
        try:
            # para cadena completa con fecha y hora
            hora_actual_dt = datetime.datetime.fromisoformat(hora_actual_dt).time()
        except ValueError:
            # para cadena solo hora
            hora_actual_dt = datetime.datetime.strptime(hora_actual_dt, "%H:%M:%S").time()
    elif isinstance(hora_actual_dt, datetime.datetime):
        hora_actual_dt = hora_actual_dt.time()
    # si ya es datetime.time, queda igual

    print(fecha_actual)

    exceptionLen = Exception.objects.filter(date=fecha_actual, time__start__lte=hora_actual_dt, time__end__gte=hora_actual_dt)

    # Obtenemos el numero total de registros a cagar
    total_exception = exceptionLen.count()

    description = f"Se cargaron {total_exception} excepciones"

    Logs.objects.create(function = "read_exception", description = description, status = "Success")

    for e in exceptionLen:

        if e.time.start <= hora_actual_dt <= e.time.end:

            alergenos_all = {
                "ALTRAMUCES": "",
                "APIO": "",
                "CACAHUETES": "",
                "CRUSTACEO": "",
                "FRUTOS_SECOS": "",
                "GLUTEN": "",
                "HUEVOS": "",
                "LACTEOS": "",
                "MOLUSCO": "",
                "MOSTAZA": "",
                "PESCADO": "",
                "SESAMO": "",
                "SOJA": "",
                "SULFITO": ""
            }

            translation_all = {
                "TRANSLATION_1": "",
                "TRANSLATION_2": "",
                "TRANSLATION_3": "",
                "TRANSLATION_4": ""
            }

            dish_id=e.dish_id
            assignment_id=e.assignment_id
            label_id=e.label_id

            # obtenemos el ID del plato
            dish=Dishes.objects.filter(id=dish_id).first()

            # Obtenemos el id del articulo a editar
            assignment=Assignment.objects.filter(id=assignment_id).first()

            # Obtenemos el plato original para agregar la excepcion
            original_dish=Dishes.objects.filter(id=assignment.dish_id).first()

            # Obtenemos las traducciones del plato
            dishlang=DishesLang.objects.filter(dish_id=dish_id).values_list('translation', flat=True)

            # Obtenemos el label del articulo a editar
            labels=Label.objects.filter(id=label_id).first()


            dish = Dishes.objects.get(id=dish_id)
            recipes = Recipe.objects.filter(dish=dish).prefetch_related('allergens')

            # Luego, modifica tu código actual así:
            alergenos = alergenos_all.copy()  # Crea una copia del diccionario base

            for recipe in recipes:
                for allergen in recipe.allergens.all():
                    if allergen.id == 0:
                        vegetarian = "YES"
                    else:
                        allCount += 1
                        alergenos[allergen.allergen.upper()] = str(allCount)
                        vegetarian = "NO"

            allCount=0


            # --------------------------------------------------------------------------
            # Consulta para obtener idiomas con estado verdadero y posición ascendente
            # --------------------------------------------------------------------------

            translations_dict = translation_all.copy()  # Crea una copia del diccionario base

            dish_id = e.dish_id

            # Paso 4: Obtener los idiomas activos ordenados por posición
            languages = Language.objects.filter(status=True).order_by('position')

            # Paso 5: Para cada idioma, obtener la traducción correspondiente del plato
            for index, language in enumerate(languages, start=1):
                translation = DishesLang.objects.filter(dish_id=dish_id, language=language.code).first()

                if translation:
                    # Agregar la traducción al diccionario con el formato adecuado
                    translations_dict[f"TRANSLATION_{index}"] = translation.translation

            # print(f"Hoy es: {e.date} y tienes una excepcion agregaste el plato {dish.dish} por el plato {original_dish.dish}")
            # print(f"ID de la asignacion: {assignment.id}, Plato original: {original_dish.id}, Category_03: {category_03_text}, Labels nomenclatura: {labels.mac}/{labels.label}, Recipe: {recipes}")
            logging.info(f"DIsh ID {dish_id}")
            logging.info(f"DIsh ID {original_dish.id}")
            logging.info(f"DIsh ID {original_dish.dish}")
            
            print(f"DIsh ID {dish_id}")
            print(f"DIsh ID {original_dish.id}")
            print(f"DIsh ID {original_dish.dish}")

            exceptionList.append({
                "articleId": f"{labels.mac}",
                "data": {
                "IS_EMPTY": "NO",
                "HAS_EXCEPTION": "YES",
                "IS_VEGETARIAN": vegetarian,
                "EXCEPTION_NAME":f"{dish.dish}",
                "EXCEPTION_DESCRIPTION":", ".join(dishlang),
                **alergenos,
                **translations_dict
                }
            })
            
    print(exceptionList)

    eupdateUrl = f"https://eu.common.solumesl.com/common/api/v2/common/articles?company={company}&store={code}"

    eupdatePayload = json.dumps(exceptionList)

    eupdateHeaders = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {access_token}'
    }

    eupdateResponse = requests.put(eupdateUrl, headers=eupdateHeaders, data=eupdatePayload)

    logging.info(f"Excepciones: {total_exception}")

    return eupdateResponse.status_code

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Actualizar producto en solum ---------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def add_update_solum(dish_id, label_id, user):

    # Configuramos el esquema y obtenemos los valores de store
    set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']
    
    logging.info(f"Plato recibido: {dish_id}")
    logging.info(f"Label recibido: {label_id}")

    allCount=0
        
    logging.info(f"El token es: {access_token}")

    alergenos_all = {
        "ALTRAMUCES": "",
        "APIO": "",
        "CACAHUETES": "",
        "CRUSTACEO": "",
        "FRUTOS_SECOS": "",
        "GLUTEN": "",
        "HUEVOS": "",
        "LACTEOS": "",
        "MOLUSCO": "",
        "MOSTAZA": "",
        "PESCADO": "",
        "SESAMO": "",
        "SOJA": "",
        "SULFITO": ""
    }

    translation_all = {
        "TRANSLATION_1": "",
        "TRANSLATION_2": "",
        "TRANSLATION_3": "",
        "TRANSLATION_4": ""
    }

    messages=[]

    # obtenemos el ID del plato
    dish=Dishes.objects.filter(id=dish_id).first()
    logging.info(f"El nombre del plato es: {dish.dish}")
    

    # Obtenemos el label del articulo a editar
    labels=Label.objects.filter(id=label_id).first()
    logging.info(f"El mac del label es es: {labels.mac}")

    dish = Dishes.objects.get(id=dish_id)

    recipes = Recipe.objects.filter(dish=dish).prefetch_related('allergens')

    alergenos = alergenos_all.copy()  # Crea una copia del diccionario base

    for recipe in recipes:
        for allergen in recipe.allergens.all():
            allCount += 1
            alergenos[allergen.allergen.upper()] = str(allCount)

    allCount=0

    # --------------------------------------------------------------------------
    # Consulta para obtener idiomas con estado verdadero y posición ascendente
    # --------------------------------------------------------------------------
    translations_dict = translation_all.copy()

    # Paso 4: Obtener los idiomas activos ordenados por posición
    languages = Language.objects.filter(status=True).order_by('position')

    # Paso 5: Para cada idioma, obtener la traducción correspondiente del plato
    for index, language in enumerate(languages, start=1):
        translation = DishesLang.objects.filter(dish_id=dish_id, language=language.code).first()

        if translation:
            # Agregar la traducción al diccionario con el formato adecuado
            translations_dict[f"TRANSLATION_{index}"] = translation.translation

    
    updateSolumUrl = f"https://eu.common.solumesl.com/common/api/v2/common/articles?company={company}&store={code}"

    if dish_id == 0:
        updateSolumPayload = json.dumps([{
            "articleId": f"{labels.mac}",
            "data": {
                "ITEM_DESCRIPTION": f"{dish.dish}",
                "IS_EMPTY": "YES",
                "HAS_EXCEPTION": "NO",
                "EXCEPTION_NAME":f"",
                "EXCEPTION_DESCRIPTION":"",
                **alergenos,
                **translations_dict
            }
        }])
    else:
        updateSolumPayload = json.dumps([{
            "articleId": f"{labels.mac}",
            "data": {
                "ITEM_DESCRIPTION": f"{dish.dish}",
                "IS_EMPTY": "NO",
                "HAS_EXCEPTION": "NO",
                "EXCEPTION_NAME":f"",
                "EXCEPTION_DESCRIPTION":"",
                **alergenos,
                **translations_dict
            }
        }])

    updateSolumHeaders = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {access_token}'
    }
    
    
    logging.info(f"Company:: {company}")
    logging.info(f"Code: {code}")
    logging.info(f"Array que se envia a solum: {updateSolumPayload}")
    
    exceptionResponse = requests.put(updateSolumUrl, headers=updateSolumHeaders, data=updateSolumPayload)
    
    logging.info(f"Respuesta de SOLUM: {exceptionResponse.status_code}, Contenido: {exceptionResponse.text}")

    if exceptionResponse.status_code == 200:

        Logs.objects.create(function = "add_update_solum", description = f"Plato actualizado correctamente en SOLUM Plato ID: {dish_id}, Label: {labels.mac}/{labels.label}", status = "Success")

        messages.append(f"Plato cargado correctamente a SOLUM Plato ID: {dish_id}, Label: {labels.mac}/{labels.label}")
    else:
        messages.append(f"No se puedo actualizar el plato en SOLUM DishID: {dish_id}, LabelsID {label_id}")
        Logs.objects.create(function = "add_update_solum", description = "No se puedo actualizar el plato en SOLUM", status = "Error")

    return messages

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Regenerate all image
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def regenerate_all_image(company, code, access_token):

    regenerateImageUrl = f"https://eu.common.solumesl.com/common/api/v2/common/labels/regenerate/images?company={company}&store={code}&regenerateType=ALL"

    regenerateImagepayload = {}

    regenerateImageHeaders = {
        'Authorization': f'Bearer {access_token}'
    }

    response = requests.request("POST", regenerateImageUrl, headers=regenerateImageHeaders, data=regenerateImagepayload)

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Mostrar % de cumplimiento de menú por rango de fechas
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def calculate_compliance(start_date, end_date, turn, user):
    try:
        set_schema(user)

        print("Iniciando calculate_compliance")
        print(f"Parámetros recibidos: start_date={start_date}, end_date={end_date}, turn={turn}")

        # Verificar si hay menús
        menus = DayliMenu.objects.filter(date__range=[start_date, end_date])
        print(f"Menús encontrados: {menus.count()}")

        if turn:
            menus = menus.filter(turn=turn)
            print(f"Menús con turno {turn}: {menus.count()}")

        # Imprimir los menús para verificar
        for menu in menus:
            print(f"Menú: {menu.date} - {menu.turn}")

        cumplimiento_expr = ExpressionWrapper(
            100 - (F('exception') * 100.0 / F('assignments')),
            output_field=FloatField()
        )

        cumplimiento_total = menus.aggregate(
            cumplimiento_avg=Avg(cumplimiento_expr)
        )['cumplimiento_avg']
        print(f"Cumplimiento total: {cumplimiento_total}")

        # Verificar excepciones
        exceptions_with_dish = Exception.objects.filter(
            date__range=(start_date, end_date)
        )

        if turn:
            exceptions_with_dish = exceptions_with_dish.filter(time__name=turn)

        print(f"Excepciones encontradas: {exceptions_with_dish.count()}")

        platos_excepciones = []
        dish_exceptions_count = exceptions_with_dish.values(
            'assignment__dish__id',
            'assignment__dish__dish'
        ).annotate(
            count=Count('id')
        ).order_by('-count')

        for dish in dish_exceptions_count:
            platos_excepciones.append({
                'dish_id': dish['assignment__dish__id'],
                'dish_name': dish['assignment__dish__dish'],
                'count': dish['count'],
                'turno': turn if turn else 'Todos',
                'fecha_inicio': start_date,
                'fecha_fin': end_date
            })

        result = {
            'cumplimiento_total': cumplimiento_total if cumplimiento_total is not None else 0,
            'platos_excepciones': platos_excepciones
        }

        print(f"Resultado final: {result}")
        return result

    except Exception as e:
        print(f"Error en calculate_compliance: {str(e)}")
        return {
            'cumplimiento_total': 0,
            'platos_excepciones': [],
            'error': str(e)
        }

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Mostrar % de cumplimiento de menú por rango de fechas
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def get_dish_exceptions(start_date, end_date, turn, user, page=1, page_size=10, search=None, ordering=None):
    set_schema(user)
    
    if isinstance(start_date, str):
        start_date = timezone.datetime.strptime(start_date, "%Y-%m-%d").date()
    if isinstance(end_date, str):
        end_date = timezone.datetime.strptime(end_date, "%Y-%m-%d").date()

    base_filter = Q(exception__date__range=[start_date, end_date])
    if turn:
        base_filter &= Q(exception__time__name=turn)

    exceptions_prefetch = Prefetch(
        'exception_set',
        queryset=Exception.objects.filter(date__range=[start_date, end_date])
        .select_related('dish', 'week', 'time', 'label'),
        to_attr='filtered_exceptions'
    )

    assignments = Assignment.objects.annotate(
        exception_count=Count('exception', filter=base_filter)
    ).filter(exception_count__gt=0)

    # Aplicar filtro de búsqueda solo para a.dish.dish
    if search:
        assignments = assignments.filter(dish__dish__icontains=search)

    # Aplicar ordenación
    if ordering:
        assignments = assignments.order_by(ordering)
    else:
        assignments = assignments.order_by('-exception_count')

    assignments = assignments.select_related('dish', 'week', 'time').prefetch_related(exceptions_prefetch)

    paginator = Paginator(assignments, page_size)
    page_obj = paginator.get_page(page)

    return [
        {
            "assignment_id": a.id,
            "original_dish": a.dish.dish,
            "week": a.week.name,
            "time": a.time.name,
            "exception_count": a.exception_count,
            "exceptions": [
                {
                    "exception_dish": e.dish.dish,
                    "date": e.date,
                    "week": e.week.name,
                    "time": e.time.name,
                    "label": e.label.label if e.label else None
                } for e in a.filtered_exceptions
            ]
        } for a in page_obj
    ], paginator.count

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Gateways
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def solum_gateway(user):
    
    # Configuramos el esquema y obtenemos los valores de store
    set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']

    url = f"https://eu.common.solumesl.com/common/api/v2/common/gateway?company={company}&store={code}&unassignOnlineLabelEnabled=false"

    payload = {}
    headers = {
    'Authorization': f'Bearer {access_token}'
    }

    response = requests.get(url, headers=headers, data=payload)

    return response.json()

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Labels Regenerate
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def labels_regenerate(user):

    # Configuramos el esquema y obtenemos los valores de store
    set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']

    url = f"https://eu.common.solumesl.com/common/api/v2/common/labels/regenerate/images?company={company}&store={code}&regenerateType=ALL"

    payload = {}
    headers = {
    'Authorization': f'Bearer {access_token}'
    }

    response = requests.post(url, headers=headers, data=payload)

    return response.json()

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Patch Gateways
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def patch_request(solum_gateway, user):

    # Configuramos el esquema y obtenemos los valores de store
    set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']

    url = f"https://eu.common.solumesl.com/common/api/v2/common/gateway?company={company}&store={code}&gateway={solum_gateway}"

    payload = {}
    headers = {
    'Authorization': f'Bearer {access_token}'
    }

    response = requests.patch(url, headers=headers, data=payload)

    return response.json()

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Envio de emails
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def send_email(sender_email, title_email, mensaje_email, hotel_name):
    # El destinatario siempre será soporte@sincroni.io
    receiver_email = "soporte@sincroni.io"
    password = "kdaqeltazmzbdcpw"  # Asegúrate de manejar la contraseña de forma segura

    # Crear el cuerpo del mensaje incluyendo el nombre del hotel
    body = f"Mensaje de {sender_email} \n\nHotel:{hotel_name} \n\n{mensaje_email}"

    msg = MIMEMultipart()
    msg['Subject'] = title_email
    msg['From'] = sender_email  # Cambia aquí para usar el sender_email
    msg['To'] = receiver_email
    msg['Reply-To'] = sender_email  # Aquí se establece el correo del remitente
    msg.attach(MIMEText(body))  # Usar el cuerpo modificado

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login("soporte@sincroni.io", password)  # Inicia sesión con la cuenta de soporte
            server.sendmail(sender_email, receiver_email, msg.as_string())  # Envía el correo
            return "Email sent successfully!"
    except smtplib.SMTPException as e:
        return f"Error: {e}"

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Crear Organizacion
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def create_organization(user, organization, company, store_code):
    print("Los datos para crear la organización en el esquema son:")
    print(user)
    print(organization)
    print(company)
    print(store_code)
    
    # Obtener el esquema del usuario
    schema = UserApp.objects.filter(username=user).first()
    if not schema:
        raise ValueError(f"No se encontró un esquema asociado al usuario {user}")
    
    select_schema = schema.schema
    print(f"Esquema seleccionado: {select_schema}")
    
    # Cambiar temporalmente el esquema
    with connection.cursor() as cursor:
        cursor.execute(f"SET search_path TO {select_schema}")
        print(f"Esquema activo: {select_schema}")
        
        # Crear la organización en el esquema seleccionado
        Organization.objects.create(name=organization, company=company, store_code=store_code)
        
        # Restaurar esquema (opcional, ya que se restablece al final de la conexión)
        cursor.execute("SET search_path TO public")
        print("Esquema restaurado a public")

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# LIsta de codigos para traduccion
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def get_deepl_target_languages():

    url = "https://api.deepl.com/v2/languages"
    params = {"type": "target"}
    headers = {
        "Authorization": f"DeepL-Auth-Key 5c16caf8-8b42-4789-83b3-25aca76925f7",
        "User-Agent": "YourApp/1.2.3"
    }
    
    response = requests.get(url, params=params, headers=headers)
    if response.status_code == 200:
        return response.json()  # Retorna la lista de idiomas
    else:
        # Si la respuesta no es 200, se levanta una excepción con el error
        response.raise_for_status()
        
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Crear traducciones preview de un plato
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def create_translations(user, dish):

    set_schema(user)
    
    # Obtiene el último ID de Dishes
    last_dish_id = Dishes.objects.last()
    
    if last_dish_id is not None:
        next_id = last_dish_id.id + 1
    else:
        next_id = 1
        
    auth_key = '5c16caf8-8b42-4789-83b3-25aca76925f7'  # Reemplaza con tu clave de autenticación
    translator = deepl.Translator(auth_key)
        
    # Filtra los idiomas activos en la base de datos
    active_languages = Language.objects.filter(status=True)
    inactive_languages = Language.objects.filter(status=False).first()

    # Lista para almacenar las instancias de DishesLang
    dishes_lang_instances = []

    # Realiza las traducciones
    for l in active_languages:
        dishTranslate = translator.translate_text(dish, target_lang=l.code, source_lang=inactive_languages.code)
        print(f"Code: {l.code} Plato a traducir: {dish} Traducción: {dishTranslate.text}")
            
        dishes_lang_instances.append({
            "id": next_id,
            "Dish": dish,
            "Code": l.code,
            "Translate": dishTranslate.text
        })
            
    return json.dumps(dishes_lang_instances)


# import os
# import json
# import openai

# from your_app.models import Language, Dishes  # ajusta el import a tu proyecto

# def create_translations(user, dish):
#     # 1. Configura schema (si usas multitenancy)
#     set_schema(user)
    
#     # 2. Calcula el próximo ID de DishesLang
#     last = Dishes.objects.last()
#     next_id = last.id + 1 if last else 1

#     # 3. Prepara OpenAI
#     openai.api_key = os.getenv("OPENAI_API_KEY")  # o pon tu clave directamente
#     model = "gpt-4"  # o el modelo que prefieras
    
#     # 4. Obtiene idiomas activos e inactivo (origen)
#     active_langs = Language.objects.filter(status=True)
#     source_lang = Language.objects.filter(status=False).first()
#     source_code = source_lang.code if source_lang else "EN"  # fallback

#     results = []

#     # 5. Traduce plato a cada idioma activo
#     for lang in active_langs:
#         target_code = lang.code.upper()
        
#         # Construye prompt semántico
#         prompt = (
#             f"Eres un sistema de traducción especializado en nombres de platos de menú. "
#             f"Traduce el nombre del plato siguiente del {source_code} al {target_code}, "
#             f"manteniendo la naturalidad y adecuándolo al estilo gastronómico:\n\n"
#             f"### Plato:\n{dish}\n\n"
#             f"### Traducción:"
#         )

#         resp = openai.ChatCompletion.create(
#             model=model,
#             messages=[
#                 {"role": "system", "content": "Eres un traductor culinario experto."},
#                 {"role": "user",   "content": prompt}
#             ],
#             temperature=0.2,
#             max_tokens=60
#         )
#         translation = resp.choices[0].message.content.strip()

#         # Debug
#         print(f"[{target_code}] \"{dish}\" → \"{translation}\"")

#         results.append({
#             "id": next_id,
#             "Dish": dish,
#             "Code": target_code,
#             "Translate": translation
#         })

#     return json.dumps(results, ensure_ascii=False, indent=2)

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def save_translations(user, dish_name, code_1=None, code_2=None, code_3=None, code_4=None,
                      trans_1=None, trans_2=None, trans_3=None, trans_4=None, allergen_ids=None,
                      all_schemas_is_verified=False):
    
    # Obtiene la información del schema
    schema_name, schema_id, schema_role = set_schema(user)
    
    if schema_name.startswith('bs_') and all_schemas_is_verified:
        # Si se debe aplicar a todos los schemas que comienzan con "bs_"
        for schema in settings.SCHEMAS:
            if schema.startswith('bs_'):
                with connection.cursor() as cursor:
                    cursor.execute(f"SET search_path TO {schema};")
                    
                    print(schema)
                    
                    # Determina el siguiente ID para el plato
                    last_dish = Dishes.objects.last()
                    next_id = last_dish.id + 1 if last_dish else 1
                    
                    # Crea el plato
                    dish = Dishes.objects.create(id=next_id, dish=dish_name.upper())
                    
                    # Construye la lista de instancias de DishesLang verificando cada par
                    dishes_lang_instances = []
                    if code_1 and trans_1:
                        dishes_lang_instances.append(
                            DishesLang(dish=dish, language=trans_1, translation=code_1.upper())
                        )
                    if code_2 and trans_2:
                        dishes_lang_instances.append(
                            DishesLang(dish=dish, language=trans_2, translation=code_2.upper())
                        )
                    if code_3 and trans_3:
                        dishes_lang_instances.append(
                            DishesLang(dish=dish, language=trans_3, translation=code_3.upper())
                        )
                    if code_4 and trans_4:
                        dishes_lang_instances.append(
                            DishesLang(dish=dish, language=trans_4, translation=code_4.upper())
                        )
                    
                    # Inserta las traducciones de forma masiva
                    with transaction.atomic():
                        DishesLang.objects.bulk_create(dishes_lang_instances)
                    
                    # Asocia alérgenos si se enviaron
                    if allergen_ids:
                        recipe = Recipe.objects.create(dish=dish)
                        allergens = Allergens.objects.filter(id__in=allergen_ids)
                        recipe.allergens.set(allergens)
                    else:
                        recipe = Recipe.objects.create(dish=dish)
                        recipe.allergens.set([]) 
                        
        response_data = {"message": "Translations saved successfully."}
    
    else:
        # Proceso para un único schema
        dishes_lang_instances = []
        last_dish = Dishes.objects.last()
        next_id = last_dish.id + 1 if last_dish else 1
        
        # Crea el plato
        dish = Dishes.objects.create(id=next_id, dish=dish_name.upper())
        
        # Verifica y crea las instancias de DishesLang para cada par recibido
        if code_1 and trans_1:
            dishes_lang_instances.append(
                DishesLang(dish=dish, language=trans_1, translation=code_1.upper())
            )
        if code_2 and trans_2:
            dishes_lang_instances.append(
                DishesLang(dish=dish, language=trans_2, translation=code_2.upper())
            )
        if code_3 and trans_3:
            dishes_lang_instances.append(
                DishesLang(dish=dish, language=trans_3, translation=code_3.upper())
            )
        if code_4 and trans_4:
            dishes_lang_instances.append(
                DishesLang(dish=dish, language=trans_4, translation=code_4.upper())
        )
        
        # Inserta las traducciones de forma masiva
        with transaction.atomic():
            DishesLang.objects.bulk_create(dishes_lang_instances)
        
        # Asocia alérgenos si se enviaron
        if allergen_ids:
            recipe = Recipe.objects.create(dish=dish)
            allergens = Allergens.objects.filter(id__in=allergen_ids)
            recipe.allergens.set(allergens)
        else:
            recipe = Recipe.objects.create(dish=dish)
            recipe.allergens.set([]) 
        
        response_data = {
            "message": "Translations saved successfully.",
            "id": next_id,
            "translations": [instance.translation for instance in dishes_lang_instances]
        }
    
    return json.dumps(response_data)

# --------------------------------------------------------------------------------------------------------------------------------------------------------------      
# listado de todos los esquemas
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def list_schemas(user):
    hotels = []

    schema_name, schema_id, schema_role = set_schema(user)

    with schema_context('public'):
        hotel = user_apps = UserApp.objects.filter(role__admin=True).distinct('username') 

    # Verificar si se encontraron registros
    if not hotel.exists():
        hotels.append({"message": "No se encontraron registros para el esquema proporcionado."})
    else:
        for u in hotel:
            hotels.append({
                "Hotel_name": u.full_name,
                "Hotel_schema": u.schema,
                "HotelID": u.id
            })

    # Devolver la respuesta como un JSON
    return json.dumps(hotels)
        
# --------------------------------------------------------------------------------------------------------------------------------------------------------------      
# listado de hoteles administrados
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def list_hotel(user):
    hotels = []
    organization=None
    print(f"Mi usuario autenticado es: {user}")

    schema_name, schema_id, schema_role = set_schema(user)

    print(f"El esquema del hotel seleccionado es: {schema_name}")
    
    # Usar 'elif' para evitar que múltiples condiciones se ejecuten
    if schema_name.startswith('bs_'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'bs_'
            if schema_role == 1 or schema_role == 4: 
                hotel = UserApp.objects.filter(role__limited=True, schema__startswith='bs_').select_related('organization').distinct('organization__name')
            else:
                hotel = UserApp.objects.filter(id=schema_id, schema__startswith='bs_').select_related('organization').distinct('organization__name')
                
        
    elif schema_name.startswith('bf_'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'bf_'
            if schema_role == 1 or schema_role == 4: 
                hotel = UserApp.objects.filter(role__limited=True, schema__startswith='bf_').select_related('organization').distinct('organization__name')
            else:
                hotel = UserApp.objects.filter(id=schema_id, schema__startswith='bf_').select_related('organization').distinct('organization__name')
    
    elif schema_name.startswith('kp_'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'kp_'
            if schema_role == 1 or schema_role == 4: 
                hotel = UserApp.objects.filter(role__in=[1, 2, 3, 4], schema__startswith='kp_').select_related('organization').distinct('organization__name')
            else:
                hotel = UserApp.objects.filter(id=schema_id, schema__startswith='kp_').select_related('organization').distinct('organization__name')
    
    elif schema_name.startswith('lp_'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'lp_'
            if schema_role == 1 or schema_role == 4: 
                hotel = UserApp.objects.filter(role__in=[1, 2, 3, 4], schema__startswith='lp_').select_related('organization').distinct('organization__name')
            else:
                hotel = UserApp.objects.filter(id=schema_id, schema__startswith='lp_').select_related('organization').distinct('organization__name')
                
    elif schema_name.startswith('monu'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'lp_'
            if schema_role == 1 or schema_role == 4: 
                hotel = UserApp.objects.filter(role__in=[1, 2, 3, 4], schema__startswith='monu').select_related('organization').distinct('organization__name')
            else:
                hotel = UserApp.objects.filter(id=schema_id, schema__startswith='monu').select_related('organization').distinct('organization__name')

    # Verificar si se encontraron registros
    if not hotel.exists():
        hotels.append({"message": "No se encontraron registros para el esquema proporcionado."})
    else:
        for u in hotel:
            hotels.append({
                "HotelID": u.id,
                "Hotel": u.organization.name,
                "Hotel_schema": u.schema
            })

    # Devolver la respuesta como un JSON
    return json.dumps(hotels)

def user_list(user):
    
    users = []

    schema_name, schema_id, schema_role = set_schema(user)
    
    # Usar 'elif' para evitar que múltiples condiciones se ejecuten
    if schema_name.startswith('bs_'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'bs_'
            if schema_role != 1:
                user_list=UserApp.objects.exclude(role_id=1).filter(schema__startswith='bs_')
            else:
                user_list=UserApp.objects.filter(schema__startswith='bs_')
                    
    elif schema_name.startswith('bf_'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'bf_'  
            if schema_role != 1:
                user_list=UserApp.objects.exclude(role_id=1).filter(schema__startswith='bf_')
            else:
                user_list=UserApp.objects.filter(schema__startswith='bf_')
    
    elif schema_name.startswith('kp_'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'kp_'
            if schema_role != 1:
                user_list=UserApp.objects.exclude(role_id=1).filter(schema__startswith='kp_')
            else:
                user_list=UserApp.objects.filter(schema__startswith='kp_')
    
    elif schema_name.startswith('lp_'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'lp_'
            if schema_role != 1:
                user_list=UserApp.objects.exclude(role_id=1).filter(schema__startswith='lp_')
            else:
                user_list=UserApp.objects.filter(schema__startswith='lp_')
    
    elif schema_name.startswith('monu'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'lp_'
            if schema_role != 1:
                user_list=UserApp.objects.exclude(role_id=1).filter(schema__startswith='monu')
            else:
                user_list=UserApp.objects.filter(schema__startswith='monu')

    # Verificar si se encontraron registros
    if not user_list.exists():
        users.append({"message": "No se encontraron registros para el esquema proporcionado."})
    else:
        for u in user_list:
            
            r=Role.objects.filter(id=u.role_id).first()
            o=Organization.objects.filter(id=u.organization_id).first()
            
            if r:
                role = {
                    "id": r.id,
                    "limited": r.limited,
                    "corporated": r.corporated,
                    "advanced": r.advanced,
                    "admin": r.admin
                }
            else:
                role= None
                
            if o:
                organization = {
                    "id": o.id,
                    "name": o.name,
                    "company": o.company,
                    "store_code": o.store_code
                }
            else:
                organization= None
            
            users.append({
                "id": u.id,
                "role": role,
                "organization": organization,
                "username": u.username,
                "full_name": u.full_name,
                "enabled": u.enabled,
                "schema": u.schema, # Aquí se agrega el role con todas sus columnas
            })

    # Devolver la respuesta como un JSON
    return json.dumps(users)    

def schemas_list(user):
    
    schema_list = []

    schema_name, schema_id, schema_role = set_schema(user)
    
    if schema_name.startswith('bs_'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'bs_'
            if schema_role != 1:
                schema_select = UserApp.objects.exclude(role_id=1).filter(schema__startswith='bs_').distinct('schema')
            else:
                schema_select = UserApp.objects.all().distinct('schema')
                
    elif schema_name.startswith('bf_'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'bs_'
            if schema_role != 1:
                schema_select = UserApp.objects.exclude(role_id=1).filter(schema__startswith='bf_').distinct('schema')
            else:
                schema_select = UserApp.objects.all().distinct('schema')
                
    elif schema_name.startswith('kp_'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'bs_'
            if schema_role != 1:
                schema_select = UserApp.objects.exclude(role_id=1).filter(schema__startswith='kp_').distinct('schema')
            else:
                schema_select = UserApp.objects.all().distinct('schema')
                
    elif schema_name.startswith('lp_'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'bs_'
            if schema_role != 1:
                schema_select = UserApp.objects.exclude(role_id=1).filter(schema__startswith='lp_').distinct('schema')
            else:
                schema_select = UserApp.objects.all().distinct('schema')
                
    elif schema_name.startswith('monu'):
        with schema_context('public'):
            # Filtrar registros donde 'schema' comienza con 'bs_'
            if schema_role != 1:
                schema_select = UserApp.objects.exclude(role_id=1).filter(schema__startswith='monu').distinct('schema')
            else:
                schema_select = UserApp.objects.all().distinct('schema')
        
    # Verificar si se encontraron registros
    if not schema_select.exists():
        schema_list.append({"message": "No existen esquemas registrados."})
    else:
        for s in schema_select:
            with schema_context('public'):
                organization_select = Organization.objects.filter(id=s.organization_id).first()
            
            schema_list.append({
                "id": organization_select.id,
                "organization": organization_select.name,
                "company": organization_select.company,
                "store_code": organization_select.store_code,
                "schema": s.schema,
            })

    # Devolver la respuesta como un JSON
    return json.dumps(schema_list)
        
# --------------------------------------------------------------------------------------------------------------------------------------------------------------      
# Cambiar de esquema usuario para varios hoteles
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def change_hotel(user, userID):
    message = []
    
    with schema_context('public'):
        adminID=UserApp.objects.filter(id=userID).first()
        
        username = adminID.username
        
        select_schema = adminID.schema
        
        print(f"El usuario en sesion es: {user}")
        
        print(f"El usuario seleccionado es: {username}")
        
        print(f"EL esquema seleccionado segun el id es: {select_schema}")
            
        print(f"Esto es lo que se obtuvo de la consulta con la variable usr: {username}")
        
        usr=UserApp.objects.filter(username=user).first()
        
        print(f"Se encontraron coincidencias de usuarios el usuario en sesion es igual al usuario de userapp {usr.username}")
        
        UserApp.objects.filter(username=usr).update(schema=select_schema)
            
        if select_schema:
            message.append({
                "user": adminID.username,
                "schema": select_schema,
                "userID": adminID.id,
            })
        else:
            message.append({
                "user": adminID.username,
                "schema": "public",
                "userID": adminID.id,
            })
        
    return json.dumps(message)

# --------------------------------------------------------------------------------------------------------------------------------------------------------------      
# Exportar menu a Excel
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def export_assignments_to_excel(user, week_id):
    # 1) Configurar esquema
    set_schema(user)

    # 2) Mapa de días
    days_map = {
        1: "Lunes", 2: "Martes", 3: "Miércoles",
        4: "Jueves", 5: "Viernes", 6: "Sábado", 7: "Domingo"
    }

    # 3) Obtener asignaciones
    assignments = (
        Assignment.objects
        .select_related('dish', 'week', 'label', 'time')
        .filter(week_id=week_id)
    )

    # 4) Estructurar datos por turno y etiqueta
    structured_data = {"Desayuno": [], "Almuerzo": [], "Cena": []}
    for asg in assignments:
        turno = asg.time.name
        etiqueta = asg.label.label if asg.label else "Sin etiqueta"
        dia = days_map.get(asg.day_of_week, "")
        plato = asg.dish.dish

        if turno in structured_data:
            idx = next(
                (i for i, row in enumerate(structured_data[turno])
                 if row["Etiqueta"] == etiqueta),
                None
            )
            if idx is None:
                new_row = {"Etiqueta": etiqueta, **{d: "" for d in days_map.values()}}
                new_row[dia] = plato
                structured_data[turno].append(new_row)
            else:
                structured_data[turno][idx][dia] += f"\n{plato}"

    # 5) Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="assignments.xlsx"'

    # 6) Crear el Excel con pandas + openpyxl
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        full_df = pd.DataFrame()

        for turno in ["Desayuno", "Almuerzo", "Cena"]:
            data = structured_data[turno]
            if not data:
                continue

            # Fila en blanco, título de sección, datos y otra fila en blanco
            blank = pd.DataFrame(
                {"Etiqueta": [""], **{d: [""] for d in days_map.values()}}
            )
            title = pd.DataFrame(
                {"Etiqueta": [turno.upper()], **{d: [""] for d in days_map.values()}}
            )
            df_data = pd.DataFrame(data)
            full_df = pd.concat(
                [full_df, blank, title, df_data, blank],
                ignore_index=True
            )

        # 7) Escribir hoja
        full_df.to_excel(writer, index=False, sheet_name="Asignaciones")
        workbook = writer.book
        worksheet = writer.sheets["Asignaciones"]

        # 8) Aplicar fondo amarillo a las filas de sección
        yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            max_col=worksheet.max_column
        ):
            val = row[0].value
            if isinstance(val, str) and val.strip() in ["DESAYUNO", "ALMUERZO", "CENA"]:
                for cell in row:
                    cell.fill = yellow_fill


    return response

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------- Funciones para correr las ruedas mediante un crom ------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def obtener_datos_dishes(username, password, restID=None):
    # URL para obtener el token
    url_token = "https://api.app.sincroni.io/api/token/"

    # Datos para la autenticación
    payload_token = json.dumps({
        "username": username,
        "password": password
    })

    headers_token = {
        'Content-Type': 'application/json'
    }

    # Solicitar el token
    response_token = requests.post(url_token, headers=headers_token, data=payload_token)

    if response_token.status_code == 200:
        # Extraer el token de acceso
        tokens = response_token.json()
        access_token = tokens['access']
        
        print(access_token)
        
        if restID is None:
            url = "https://api.app.sincroni.io/labels_schedule/"# crear uno pasando el parametro
        else:
            url = f"https://api.app.sincroni.io/labels_schedule/{restID}"

        payload = ""
        headers = {
        'Authorization': f'Bearer {access_token}'
        }

        response = requests.request("GET", url, headers=headers, data=payload)

        return response.text
        
    else:
        print("Error al obtener el token:", response_token.text)
            
def obtener_times(username, password):
    # URL para obtener el token
    url_token = "https://api.app.sincroni.io/api/token/"
    
    print(username)
    print(password)

    # Datos para la autenticación
    payload_token = json.dumps({
        "username": username,
        "password": password
    })

    headers_token = {
        'Content-Type': 'application/json'
    }

    # Solicitar el token
    response_token = requests.post(url_token, headers=headers_token, data=payload_token)

    if response_token.status_code == 200:
        # Extraer el token de acceso
        tokens = response_token.json()
        access_token = tokens['access']
        
        print(access_token)
        
        url = "https://api.app.sincroni.io/api/times/"

        payload = ""
        headers = {
        'Authorization': f'Bearer {access_token}'
        }

        response = requests.request("GET", url, headers=headers, data=payload)

        return response.json()
        
    else:
        print("Error al obtener el token:", response_token.text)

# -------------------------------------------------------------------------------------------------------------------------------------------------------------
# traductor ---------------------------------------------------------------------------------------------------------------------------------------------------
# -------------------------------------------------------------------------------------------------------------------------------------------------------------
def traductor_masivo(user):
    
    set_schema(user)

    # Obtener todos los registros de la tabla Dishes
    dishes_queryset = Dishes.objects.all()
    dishes_list = list(dishes_queryset)
    dish_names = [dish.dish for dish in dishes_list]

    return dish_names

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def bulk_upsert_assignments(request, user):
    
    store_code(user)

    # Nombre del esquema
    schema_name, schema_id, schema_role = set_schema(user)

    products = []
    labelsUpdate = []
    messages = []

    # Se obtiene un token válido (se refresca o genera uno nuevo si fuera necesario)
    token_data = get_valid_solum_token(user)
    if not token_data:
        messages.append("No se pudo obtener un token válido")
        return messages

    # Extrae el access_token y company desde token_data
    access_token = token_data['access_token']

    labels = Label.objects.filter(~Q(label=F('mac')))

    # Obtenemos el numero total de productos y labels a cagar
    total_install = labels.count()

    payload = request.data
    if not isinstance(payload, list):
        return {"message": "Se esperaba un array JSON de asignaciones", "status": 400}

    results = []
    for item in payload:
        assignment = Assignment.objects.filter(
            week_id      = item['week'],
            label_id     = item['label'],
            time_id      = item['time'],
            day_of_week  = item['day_of_week'],
        ).first()

        if assignment:
            assignment.dish_id = item['dish_id']
            assignment.save(update_fields=['dish_id'])
            created = False
            update = True
        else:
            assignment = Assignment.objects.create(
                dish_id     = item['dish_id'],
                week_id     = item['week'],
                label_id    = item['label'],
                time_id     = item['time'],
                day_of_week = item['day_of_week'],
            )
            created = True
            update = False

        results.append({
            'id':          assignment.id,
            'dish_id':     assignment.dish_id,
            'week':        assignment.week_id,
            'label':       assignment.label_id,
            'time':        assignment.time_id,
            'day_of_week': assignment.day_of_week,
            'created':     created,
            'update':      update,
        })

    # for a in labels:
    
    #     labelsUpdate.append({f"articleId": {a.mac}, "articleName": {a.label}, "nfcUrl": "https://www.solumesl.com/p/"+{a.mac}, "eans": "[]"})
    
    # update_all_labels(user, labelsUpdate)

    return {'assignments': results, 'status': 200}

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Actualizar label en solum masivo -----------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def update_all_labels(user, labelsUpdate):
    
    # Configuramos el esquema y obtenemos los valores de store
    schema_name, schema_id, schema_role = set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']

    messages = []

    # URL para editar el artículo
    article_url = f"https://eu.common.solumesl.com/common/api/v2/common/articles?company={company}&store={code}"

    # Datos para la solicitud de actualizacion de nomenclatura
    article_payload = json.dumps(labelsUpdate)

    article_headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {access_token}'
    }
    
    # Solicitud para editar el artículo
    article_response = requests.put(article_url, headers=article_headers, data=article_payload)
    
    print(article_response.status_code)
    
    if article_response.status_code==200:
        
        messages.append(f"Se actualizaron todos los labels")
    else:
        messages.append(f"Error al actualizar los labels")
        
    return messages

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Actualizar label en solum uno a uno --------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def update_labels(user, label, mac):
    
    # Configuramos el esquema y obtenemos los valores de store
    schema_name, schema_id, schema_role = set_schema(user)
    store_code(user)  # Esto define las variables globales "company" y "code"

    # Se obtiene un token válido; get_valid_solum_token se encargará de refrescar o generar uno nuevo según convenga
    token_data = get_valid_solum_token(user)
    
    if not token_data:
        logger.error("No se pudo obtener un token válido")
        return []

    # Extraemos el access_token y el company (el store_code ya asignó la variable global "code")
    access_token = token_data['access_token']

    messages = []

    # URL para editar el artículo
    article_url = f"https://eu.common.solumesl.com/common/api/v2/common/articles?company={company}&store={code}"

    # Datos para la solicitud de actualizacion de nomenclatura
    article_payload = json.dumps([{
        "articleId": f"{mac}",
        "articleName": f"{label}",
        "nfcUrl": f"https://www.solumesl.com/p/{label}"
        }])

    article_headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {access_token}'
    }
    
    # Solicitud para editar el artículo
    article_response = requests.put(article_url, headers=article_headers, data=article_payload)
    
    print(article_response.status_code)
    
    if article_response.status_code==200:
        
        messages.append(f"Se actuali el ArticleID: {mac} con la nomenclatura: {label}")
    else:
        messages.append(f"Error al actualizar el ArticleID: {mac}")
        
    return messages

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Descargar Platos ---------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def export_dishes_excel(user):
    schema_name, schema_id, schema_role = set_schema(user)

    # 1) obtenemos los códigos de idioma activos
    with connection.cursor() as cursor:
        cursor.execute(f"""
            SELECT code
              FROM {schema_name}.dishes_language
             WHERE status = true
             ORDER BY code
        """)
        idiomas = [row[0] for row in cursor.fetchall()]

    # 2) preparamos los CASE dinámicos para traducciones
    translation_cases = ",\n    ".join(
        f"MAX(CASE WHEN dl.language = '{lang}' THEN dl.translation END) AS {lang.lower().replace('-', '_')}"
        for lang in idiomas
    )

    # 3) montamos el SQL completo con SI/NO en alérgenos
    sql = f"""
    WITH alerg AS (
      SELECT r.dish_id, dra.allergens_id
        FROM {schema_name}.dishes_recipe r
        JOIN {schema_name}.dishes_recipe_allergens dra
          ON r.id = dra.recipe_id
    )
    SELECT
        d.id   AS dish_id,
        d.dish AS nombre_español,
        {translation_cases},
        -- Alérgenos en SI/NO
        CASE WHEN MAX(CASE WHEN a.allergens_id = 0 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS vegetariano,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 1 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS altramuces,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 2 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS apio,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 3 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS cacahuetes,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 4 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS crustaceo,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 5 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS frutos_secos,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 6 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS gluten,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 7 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS huevos,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 8 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS lacteos,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 9 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS molusco,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 10 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS mostaza,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 11 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS pescado,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 12 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS sesamo,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 13 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS soja,
        CASE WHEN MAX(CASE WHEN a.allergens_id = 14 THEN 1 ELSE 0 END) = 1 THEN 'SI' ELSE 'NO' END AS sulfito
      FROM {schema_name}.dishes_dishes d
      LEFT JOIN {schema_name}.dishes_disheslang dl
        ON d.id = dl.dish_id
      LEFT JOIN alerg a
        ON d.id = a.dish_id
    GROUP BY d.id, d.dish
    ORDER BY d.dish;
    """

    # 4) ejecutamos y generamos el Excel
    with connection.cursor() as cursor:
        cursor.execute(sql)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for row in rows:
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="dishes.xlsx"'
    wb.save(response)
    return response

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Crear menu base
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def process_menu_assignments(user, data):
    # Ajustar conexión al esquema del usuario
    set_schema(user)

    # Extraer datos del payload
    restaurant_name = data.get('restaurant')
    week_start      = data.get('week_start')
    week_end        = data.get('week_end')
    time_name       = data.get('time')
    start           = data.get('start')
    end             = data.get('end')
    labels_count    = data.get('labels')

    # Validaciones básicas
    if not all([restaurant_name, time_name, start, end, labels_count is not None]):
        return {'detail': 'Faltan parámetros obligatorios.'}, status.HTTP_400_BAD_REQUEST

    try:
        labels_count = int(labels_count)
    except (TypeError, ValueError):
        return {'detail': 'El campo "labels" debe ser un entero.'}, status.HTTP_400_BAD_REQUEST

    # Crear u obtener restaurante
    restaurant_obj, _ = Restaurant.objects.get_or_create(
        name__iexact=restaurant_name,
        defaults={'name': restaurant_name, 'type': 'NULL', 'active': False}
    )

    # Crear u obtener horario
    time_obj, _ = Time.objects.update_or_create(
        name__iexact=time_name,
        start=start,
        end=end,
        defaults={
            'name':       time_name.capitalize(),
            'start':      start,
            'end':        end,
            'restaurant': restaurant_obj
        }
    )

    # Crear u obtener semana
    if week_start and week_end:
        try:
            start_date = datetime.strptime(week_start, '%Y-%m-%d').date()
            end_date   = datetime.strptime(week_end,   '%Y-%m-%d').date()
        except ValueError:
            return {'detail': 'Formato de fechas inválido. Usa YYYY-MM-DD.'}, status.HTTP_400_BAD_REQUEST

        week_obj, _ = Week.objects.get_or_create(
            restaurant=restaurant_obj,
            start_date=start_date,
            end_date=end_date,
            defaults={
                'name':      restaurant_name,
                'week_num':  1,
                'week_days': 7,
                'active':    False  # Forzar valor
            }
        )
    else:
        today = timezone.localdate()
        try:
            week_obj = Week.objects.get(
                restaurant=restaurant_obj,
                start_date__lte=today,
                end_date__gte=today
            )
        except Week.DoesNotExist:
            return {'detail': 'No se encontró la semana actual para este restaurante.'}, status.HTTP_400_BAD_REQUEST

    # Validar que no existan asignaciones ya creadas para ese week y time
    if Assignment.objects.filter(week=week_obj, time=time_obj).exists():
        return {'detail': 'Ya existen asignaciones para esa semana y ese tiempo.'}, status.HTTP_400_BAD_REQUEST

    # Validar disponibilidad de labels para clonar
    total_labels = Label.objects.count()
    if total_labels < labels_count:
        return {'detail': f'Solo hay {total_labels} labels disponibles.'}, status.HTTP_400_BAD_REQUEST

    # Clonar N labels
    originales = Label.objects.order_by('id')[:labels_count]
    clones = []
    for orig in originales:
        clones.append(Label(
            mac        = orig.mac,
            model      = orig.model,
            label      = orig.label,
            enabled    = orig.enabled,
            restaurant = restaurant_obj,
            template   = orig.template
        ))
    Label.objects.bulk_create(clones)

    # Obtener los labels recién creados del nuevo restaurante
    labels = Label.objects.filter(restaurant=restaurant_obj).order_by('-id')[:labels_count][::-1]

    # Crear asignaciones para cada día de la semana
    assignments = []
    for day in range(1, 8):  # 1 = lunes, ..., 7 = domingo
        for label in labels:
            assignments.append(
                Assignment(
                    dish_id=0,
                    week=week_obj,
                    label=label,
                    time=time_obj,
                    day_of_week=day
                )
            )

    Assignment.objects.bulk_create(assignments)

    return {
        'detail': 'Asignaciones creadas correctamente.',
        'restaurant_id': restaurant_obj.id,
        'week_id': week_obj.id,
        'time_id': time_obj.id
    }, status.HTTP_201_CREATED
    
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# Eliminar gestion
# --------------------------------------------------------------------------------------------------------------------------------------------------------------
def delete_menu(user, restaurant_id):
    
    # Ajusta la conexión al esquema correspondiente
    set_schema(user)
    
    r = Restaurant.objects.filter(id=restaurant_id).first()
    w = Week.objects.filter(restaurant_id=restaurant_id).first()
    
    if w.active == False:
        Assignment.objects.filter(week__restaurant_id=restaurant_id).delete()
        Label.objects.filter(restaurant_id=restaurant_id).delete()
        Time.objects.filter(restaurant_id=restaurant_id).delete()
        Week.objects.filter(restaurant_id=restaurant_id).delete()
        Restaurant.objects.filter(id=restaurant_id).delete()
        return {"message": f"Se ha eliminado la informacion para el restaurant {r.name}."}, status.HTTP_200_OK
    
    else:
        return {"message": f"El menu gestionado para el restaurant que desea eliminar debe estar inactivo."}, status.HTTP_200_OK